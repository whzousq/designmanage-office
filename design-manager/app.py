"""
设计项目进度管理系统 - Flask后端
技术栈: Flask + SQLite + Flask-Login
"""
import os
import sys
import io
from datetime import datetime, timedelta
from functools import wraps

from flask import Flask, request, jsonify, send_from_directory, session, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
import xlsxwriter

try:
    from pypinyin import pinyin, Style
    USE_PYPINYIN = True
except ImportError:
    USE_PYPINYIN = False

# 简化的拼音映射（当pypinyin不可用时）
SIMPLE_PINYIN = {
    # 常用姓
    '王': 'wang', '李': 'li', '张': 'zhang', '刘': 'liu', '陈': 'chen',
    '杨': 'yang', '黄': 'huang', '赵': 'zhao', '周': 'zhou', '吴': 'wu',
    '徐': 'xu', '孙': 'sun', '马': 'ma', '朱': 'zhu', '胡': 'hu',
    '林': 'lin', '何': 'he', '郭': 'guo', '罗': 'luo', '高': 'gao',
    '郑': 'zheng', '梁': 'liang', '谢': 'xie', '宋': 'song', '唐': 'tang',
    '许': 'xu', '韩': 'han', '冯': 'feng', '邓': 'deng', '曹': 'cao',
    '彭': 'peng', '曾': 'zeng', '肖': 'xiao', '田': 'tian', '董': 'dong',
    '袁': 'yuan', '潘': 'pan', '于': 'yu', '蒋': 'jiang', '蔡': 'cai',
    '余': 'yu', '杜': 'du', '叶': 'ye', '程': 'cheng', '苏': 'su',
    '魏': 'wei', '吕': 'lv', '丁': 'ding', '任': 'ren', '沈': 'shen',
    '姚': 'yao', '卢': 'lu', '姜': 'jiang', '崔': 'cui', '钟': 'zhong',
    '谭': 'tan', '陆': 'lu', '汪': 'wang', '范': 'fan', '金': 'jin',
    '石': 'shi', '廖': 'liao', '贾': 'jia', '夏': 'xia', '韦': 'wei',
    '傅': 'fu', '方': 'fang', '白': 'bai', '邹': 'zou', '孟': 'meng',
    '熊': 'xiong', '秦': 'qin', '邱': 'qiu', '侯': 'hou', '江': 'jiang',
    '尹': 'yin', '薛': 'xue', '闫': 'yan', '段': 'duan', '雷': 'lei',
    '龙': 'long', '史': 'shi', '陶': 'tao', '贺': 'he', '顾': 'gu',
    '毛': 'mao', '郝': 'hao', '龚': 'gong', '邵': 'shao', '万': 'wan',
    '钱': 'qian', '严': 'yan', '孔': 'kong', '常': 'chang', '武': 'wu',
    '乔': 'qiao', '赖': 'lai', '庞': 'pang', '樊': 'fan', '殷': 'yin',
    '施': 'shi', '洪': 'hong', '翟': 'zhai', '安': 'an', '章': 'zhang',
    '梁': 'liang', '谢': 'xie', '宋': 'song', '唐': 'tang', '邓': 'deng',
    '付': 'fu', 
    # 常用名
    '伟': 'wei', '强': 'qiang', '勇': 'yong', '军': 'jun', '辉': 'hui',
    '敏': 'min', '静': 'jing', '丽': 'li', '芳': 'fang', '娜': 'na',
    '婷': 'ting', '倩': 'qian', '宁': 'ning', '琳': 'lin', '玲': 'ling',
    '华': 'hua', '明': 'ming', '磊': 'lei', '涛': 'tao', '杰': 'jie',
    '鹏': 'peng', '飞': 'fei', '燕': 'yan', '红': 'hong', '青': 'qing',
    '平': 'ping', '波': 'bo', '刚': 'gang', '峰': 'feng', '超': 'chao',
    '洋': 'yang', '亮': 'liang', '伟': 'wei', '东': 'dong', '浩': 'hao',
    '文': 'wen', '军': 'jun', '德': 'de', '志': 'zhi', '忠': 'zhong',
    '业': 'ye', '建': 'jian', '福': 'fu', '贵': 'gui', '生': 'sheng',
    '学': 'xue', '龙': 'long', '云': 'yun', '雪': 'xue', '梅': 'mei',
    '兰': 'lan', '菊': 'ju', '桂': 'gui', '英': 'ying', '秀': 'xiu',
    '珍': 'zhen', '珠': 'zhu', '琴': 'qin', '棋': 'qi', '书': 'shu',
    '画': 'hua', '诗': 'shi', '雅': 'ya', '洁': 'jie', '婷': 'ting',
    '慧': 'hui', '颖': 'ying', '欣': 'xin', '怡': 'yi', '然': 'ran',
    '思': 'si', '雨': 'yu', '雨': 'yu', '晨': 'chen', '曦': 'xi',
    '梦': 'meng', '琪': 'qi', '瑶': 'yao', '璇': 'xuan', '瑾': 'jin',
    '萱': 'xuan', '怡': 'yi', '诺': 'nuo', '依': 'yi', '涵': 'han',
    '梓': 'zi', '萱': 'xuan', '欣': 'xin', '怡': 'yi', '诺': 'nuo',
    '轩': 'xuan', '涵': 'han', '泽': 'ze', '宇': 'yu', '浩': 'hao',
    '宸': 'chen', '睿': 'rui', '铭': 'ming', '昊': 'hao', '硕': 'shuo',
    '航': 'hang', '宇': 'yu', '辰': 'chen', '熙': 'xi', '哲': 'zhe',
    '仕': 'shi', '永': 'yong','子':'zi','艳':'yan','红':'hong','辉':'hui',
    '亚':'ya','敏':'min','娟':'juan','昊':'hao','崔':'cui','巍':'wei','磊':'lei',
    '涛':'tao','杰':'jie','阊':'chang','圣':'sheng','晓':'xiao','灿':'can','玮':'wei',
    '久':'jiu',
       # 工程师常用名后缀
    '工': 'gong', '师': 'shi', '员': 'yuan'
}


def generate_username(chinese_name):
    """根据中文姓名生成用户名
    - 2个汉字：全拼（如"王工" → "wanggong"）
    - 3个及以上汉字：姓全拼+名首字母（如"邹仕强" → "zousq"）
    """
    if not chinese_name or not isinstance(chinese_name, str):
        return 'user'

    name = chinese_name.strip()
    if len(name) == 0:
        return 'user'

    pinyin_list = []

    if USE_PYPINYIN:
        try:
            raw = pinyin(name, style=Style.NORMAL)
            if raw and len(raw) > 0:
                for item in raw:
                    if isinstance(item, list) and len(item) > 0:
                        pinyin_list.append(item[0].lower())
                    elif isinstance(item, str):
                        pinyin_list.append(item.lower())
        except:
            pass

    if not pinyin_list:
        for char in name:
            py = SIMPLE_PINYIN.get(char, '')
            if not py:
                # 如果汉字不在映射表中，使用默认首字母'x'
                py = 'x'
            pinyin_list.append(py)

    pinyin_list = [p for p in pinyin_list if p]

    if len(pinyin_list) == 0:
        return 'user'
    elif len(pinyin_list) == 1:
        return pinyin_list[0]
    elif len(pinyin_list) == 2:
        return ''.join(pinyin_list)
    else:
        surname = pinyin_list[0]
        given_name_initials = ''.join([p[0] if len(p) > 0 else '' for p in pinyin_list[1:]])
        return surname + given_name_initials

# ==================== App Config ====================
app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dpm-secret-key-change-in-production-2026')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///design_manager.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

CORS(app, supports_credentials=True)
db = SQLAlchemy(app)

# ==================== Database Models ====================

class User(db.Model):
    """用户表"""
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    real_name = db.Column(db.String(80), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='staff')  # admin, engineer, staff
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'realName': self.real_name,
            'role': self.role,
            'isActive': self.is_active,
            'createdAt': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else ''
        }


class Designer(db.Model):
    """设计人员表"""
    __tablename__ = 'designers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)  # 姓名
    name_pinyin = db.Column(db.String(100))  # 姓名拼音
    department = db.Column(db.String(100))  # 所属部门
    major = db.Column(db.String(50))  # 专业（建筑/结构/给排水/暖通/电气等）
    title = db.Column(db.String(50))  # 职称
    phone = db.Column(db.String(20))  # 联系电话
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'namePinyin': self.name_pinyin or '',
            'department': self.department,
            'major': self.major,
            'title': self.title,
            'phone': self.phone,
            'isActive': self.is_active,
            'createdAt': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else ''
        }


def generate_name_pinyin(name):
    """根据姓名生成拼音
    - 2个汉字：汉语拼音全拼
    - 3个及以上汉字：姓全拼 + 后续名字的拼音首字母
    - 全部小写
    """
    if not name or USE_PYPINYIN is False:
        return ''
    
    # 移除非汉字字符
    chinese_chars = [c for c in name if '\u4e00' <= c <= '\u9fff']
    if not chinese_chars:
        return ''
    
    try:
        # 仅对汉字字符生成拼音
        chinese_text = ''.join(chinese_chars)
        pinyin_list = pinyin(chinese_text, style=Style.NORMAL)
        full_pinyin = ''.join([p[0] if p else '' for p in pinyin_list])
        
        if len(chinese_chars) == 2:
            # 2个字：全拼
            return full_pinyin.lower()
        else:
            # 3个字及以上：姓全拼 + 名字首字母
            surname = pinyin_list[0][0] if pinyin_list and pinyin_list[0] else ''
            initials = ''.join([p[0][0] if p and p[0] else '' for p in pinyin_list[1:]])
            return (surname + initials).lower()
    except:
        return ''


class ConstructionPlan(db.Model):
    """施工图计划表"""
    __tablename__ = 'construction_plans'
    id = db.Column(db.Integer, primary_key=True)
    project_name = db.Column(db.String(200), nullable=False)
    project_manager = db.Column(db.String(50))  # 项目负责人（移动到项目名称后）
    unit_project = db.Column(db.String(200))
    drawing_content = db.Column(db.String(500))  # 成品套图（原单位工程下专业图纸内容）
    designer = db.Column(db.String(50))
    major = db.Column(db.String(50))  # 专业（原专业类别）
    major_director = db.Column(db.String(50))  # 专业所长
    estimated_quantitative = db.Column(db.String(100))  # 预估量化指标
    a1_drawing_count = db.Column(db.String(50))  # 折合A1图纸张数
    plan_completion_percent = db.Column(db.String(20))  # 计划完成百分比
    plan_work_days = db.Column(db.Integer)  # 计划工作日
    start_date = db.Column(db.String(20))  # 计划开始时间
    end_date = db.Column(db.String(20))  # 计划完成时间
    monthly_deviation = db.Column(db.String(50))  # 上月滚动偏差
    completion_status = db.Column(db.String(20))  # 完成情况
    remarks = db.Column(db.String(500))
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'projectName': self.project_name,
            'projectManager': self.project_manager,
            'unitProject': self.unit_project,
            'drawingContent': self.drawing_content,
            'designer': self.designer,
            'major': self.major,
            'majorDirector': self.major_director,
            'estimatedQuantitative': self.estimated_quantitative,
            'a1DrawingCount': self.a1_drawing_count,
            'planCompletionPercent': self.plan_completion_percent,
            'planWorkDays': self.plan_work_days,
            'startDate': self.start_date,
            'endDate': self.end_date,
            'monthlyDeviation': self.monthly_deviation,
            'completionStatus': self.completion_status,
            'remarks': self.remarks,
            'updatedAt': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else ''
        }


class PhasePlan(db.Model):
    """阶段设计计划表"""
    __tablename__ = 'phase_plans'
    id = db.Column(db.Integer, primary_key=True)
    project_name = db.Column(db.String(200), nullable=False)
    project_manager = db.Column(db.String(50))
    design_phase = db.Column(db.String(50))
    product_drawing_name = db.Column(db.String(200))
    designer = db.Column(db.String(50))
    major = db.Column(db.String(200))
    major_director = db.Column(db.String(50))
    estimated_quantitative = db.Column(db.String(100))
    a1_drawing_count = db.Column(db.String(50))
    manual_page_count = db.Column(db.String(50))
    plan_completion_percent = db.Column(db.String(20))
    plan_work_days = db.Column(db.Integer)
    plan_start_date = db.Column(db.String(20))
    plan_end_date = db.Column(db.String(20))
    monthly_deviation = db.Column(db.String(50))
    remarks = db.Column(db.String(500))
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'projectName': self.project_name,
            'projectManager': self.project_manager,
            'designPhase': self.design_phase,
            'productDrawingName': self.product_drawing_name,
            'designer': self.designer,
            'major': self.major,
            'majorDirector': self.major_director,
            'estimatedQuantitative': self.estimated_quantitative,
            'a1DrawingCount': self.a1_drawing_count,
            'manualPageCount': self.manual_page_count,
            'planCompletionPercent': self.plan_completion_percent,
            'planWorkDays': self.plan_work_days if self.plan_work_days else '',
            'planStartDate': self.plan_start_date,
            'planEndDate': self.plan_end_date,
            'monthlyDeviation': self.monthly_deviation,
            'remarks': self.remarks,
            'updatedAt': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else ''
        }


class TechnicalPlan(db.Model):
    """技术要求计划表"""
    __tablename__ = 'technical_plans'
    id = db.Column(db.Integer, primary_key=True)
    project_name = db.Column(db.String(200))
    project_manager = db.Column(db.String(50))
    design_phase = db.Column(db.String(50))
    product_drawing_name = db.Column(db.String(200), nullable=False)
    designer = db.Column(db.String(50))
    major = db.Column(db.String(200))
    major_director = db.Column(db.String(50))
    estimated_quantitative = db.Column(db.String(100))
    a1_drawing_count = db.Column(db.String(50))
    manual_page_count = db.Column(db.String(50))
    plan_completion_percent = db.Column(db.String(20))
    plan_work_days = db.Column(db.Integer)
    plan_start_date = db.Column(db.String(20))
    plan_end_date = db.Column(db.String(20))
    monthly_deviation = db.Column(db.String(50))
    remarks = db.Column(db.String(500))
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'projectName': self.project_name,
            'projectManager': self.project_manager,
            'designPhase': self.design_phase,
            'productDrawingName': self.product_drawing_name,
            'designer': self.designer,
            'major': self.major,
            'majorDirector': self.major_director,
            'estimatedQuantitative': self.estimated_quantitative,
            'a1DrawingCount': self.a1_drawing_count,
            'manualPageCount': self.manual_page_count,
            'planCompletionPercent': self.plan_completion_percent,
            'planWorkDays': self.plan_work_days if self.plan_work_days else '',
            'planStartDate': self.plan_start_date,
            'planEndDate': self.plan_end_date,
            'monthlyDeviation': self.monthly_deviation,
            'remarks': self.remarks,
            'updatedAt': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else ''
        }


class MajorCategory(db.Model):
    """专业类别表"""
    __tablename__ = 'major_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'createdAt': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else ''
        }


class ProjectName(db.Model):
    """项目名称字典表 - 用于统一项目名称"""
    __tablename__ = 'project_names'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, unique=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'createdAt': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else '',
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else ''
        }


class WorkloadStats(db.Model):
    """工作量统计表"""
    __tablename__ = 'workload_stats'
    id = db.Column(db.Integer, primary_key=True)
    month = db.Column(db.String(10), nullable=False)  # 月份 YYYY-MM
    department = db.Column(db.String(100))  # 所属单位
    name = db.Column(db.String(50), nullable=False)  # 姓名
    major = db.Column(db.String(50))  # 专业
    major_director = db.Column(db.String(50))  # 专业所长
    role = db.Column(db.String(50))  # 人员角色
    work_content = db.Column(db.String(500))  # 具体完成工作内容
    quantitative_index = db.Column(db.String(100))  # 量化指标
    project_name = db.Column(db.String(200), nullable=False)  # 项目名称
    project_manager = db.Column(db.String(50))  # 项目负责人
    design_phase = db.Column(db.String(50))  # 设计阶段
    drawing_name = db.Column(db.String(200))  # 成品套图/文件名称
    a1_drawing_count = db.Column(db.String(50))  # 折合A1图纸张数
    manual_page_count = db.Column(db.String(50))  # 说明书页数
    completion_rate = db.Column(db.String(20))  # 套图文件完成率
    actual_work_days = db.Column(db.Float)  # 实际工作日
    work_start_date = db.Column(db.String(20))  # 工作开始时间
    work_end_date = db.Column(db.String(20))  # 工作结束时间
    plan_deviation = db.Column(db.Float)  # 计划偏差(天)
    remarks = db.Column(db.String(500))  # 备注
    confirmed = db.Column(db.Boolean, default=False)  # 实际工作日是否已确认
    confirmed_by = db.Column(db.Integer, db.ForeignKey('users.id'))  # 确认人
    confirmed_at = db.Column(db.DateTime)  # 确认时间
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))  # 创建人
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'month': self.month,
            'department': self.department,
            'name': self.name,
            'major': self.major,
            'majorDirector': self.major_director,
            'role': self.role,
            'workContent': self.work_content,
            'quantitativeIndex': self.quantitative_index,
            'projectName': self.project_name,
            'projectManager': self.project_manager,
            'designPhase': self.design_phase,
            'drawingName': self.drawing_name,
            'a1DrawingCount': self.a1_drawing_count,
            'manualPageCount': self.manual_page_count,
            'completionRate': self.completion_rate,
            'actualWorkDays': self.actual_work_days if self.actual_work_days else '',
            'workStartDate': self.work_start_date,
            'workEndDate': self.work_end_date,
            'planDeviation': self.plan_deviation if self.plan_deviation else '',
            'remarks': self.remarks,
            'confirmed': self.confirmed,
            'confirmedBy': self.confirmed_by,
            'confirmedAt': self.confirmed_at.strftime('%Y-%m-%d %H:%M:%S') if self.confirmed_at else '',
            'createdBy': self.created_by,
            'updatedAt': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else ''
        }


# ==================== Auth Helpers ====================

def login_required(f):
    """登录验证装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'code': 401, 'msg': '请先登录'}), 401
        return f(*args, **kwargs)
    return decorated


def edit_required(f):
    """编辑权限验证装饰器（admin或engineer）"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'code': 401, 'msg': '请先登录'}), 401
        user = db.session.get(User, session['user_id'])
        if not user or user.role == 'staff':
            return jsonify({'code': 403, 'msg': '权限不足，仅首席工程师及以上可操作'}), 403
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """管理员权限验证装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'code': 401, 'msg': '请先登录'}), 401
        user = db.session.get(User, session['user_id'])
        if not user or user.role != 'admin':
            return jsonify({'code': 403, 'msg': '仅管理员可操作'}), 403
        return f(*args, **kwargs)
    return decorated


def get_current_user():
    """获取当前登录用户"""
    if 'user_id' in session:
        return db.session.get(User, session['user_id'])
    return None


# ==================== Init Default Data ====================

def init_default_data():
    """初始化默认用户和示例数据"""
    # Create default users
    defaults = [
        {'username': 'admin', 'password': 'admin123', 'real_name': '系统管理员', 'role': 'admin'},
        {'username': 'engineer', 'password': 'eng123', 'real_name': '张工（首席工程师）', 'role': 'engineer'},
        {'username': 'staff', 'password': 'staff123', 'real_name': '李明（设计人员）', 'role': 'staff'},
    ]
    for d in defaults:
        if not User.query.filter_by(username=d['username']).first():
            user = User(username=d['username'], real_name=d['real_name'], role=d['role'])
            user.set_password(d['password'])
            db.session.add(user)
    db.session.commit()

    # Create default designers
    if Designer.query.count() == 0:
        default_designers = [
            Designer(name='王工', name_pinyin='wanggong', department='设计一部', major='建筑', title='工程师'),
            Designer(name='李工', name_pinyin='ligong', department='设计一部', major='结构', title='工程师'),
            Designer(name='赵工', name_pinyin='zhaogong', department='设计一部', major='给排水', title='高级工程师'),
            Designer(name='陈工', name_pinyin='chegong', department='设计二部', major='暖通', title='工程师'),
            Designer(name='孙工', name_pinyin='sungong', department='设计二部', major='电气', title='工程师'),
            Designer(name='刘工', name_pinyin='liugong', department='设计二部', major='总图', title='高级工程师'),
            Designer(name='张工', name_pinyin='zhanggong', department='设计一部', major='建筑', title='高级工程师'),
        ]
        db.session.add_all(default_designers)
        db.session.commit()

    # Create default major categories
    if MajorCategory.query.count() == 0:
        default_majors = [
            MajorCategory(name='建筑'),
            MajorCategory(name='结构'),
            MajorCategory(name='给排水'),
            MajorCategory(name='暖通'),
            MajorCategory(name='电气'),
            MajorCategory(name='总图'),
            MajorCategory(name='动力'),
            MajorCategory(name='通信'),
            MajorCategory(name='其他'),
        ]
        db.session.add_all(default_majors)
        db.session.commit()

    # 检查三个计划表是否已有数据，如果有任何一个表有数据，则跳过所有示例数据
    has_existing_data = (ConstructionPlan.query.count() > 0 or 
                        PhasePlan.query.count() > 0 or 
                        TechnicalPlan.query.count() > 0)
    
    if has_existing_data:
        return

    # Create sample data for construction plans
    if ConstructionPlan.query.count() == 0:
        samples = [
            ConstructionPlan(project_name='XX污水处理厂扩建工程', unit_project='生化处理车间',
                           drawing_content='建筑平立剖面图', major='建筑', designer='王工',
                           major_director='王总', estimated_quantitative='建筑专业全套施工图',
                           a1_drawing_count='120', plan_completion_percent='60%',
                           plan_work_days=121, completion_status='进行中', start_date='2026-03-01', end_date='2026-06-30',
                           project_manager='张工', remarks='一期扩建'),
            ConstructionPlan(project_name='XX污水处理厂扩建工程', unit_project='生化处理车间',
                           drawing_content='结构施工图', major='结构', designer='李工',
                           major_director='李总', estimated_quantitative='结构专业全套施工图',
                           a1_drawing_count='150', plan_completion_percent='30%',
                           plan_work_days=106, completion_status='未开始', start_date='2026-04-01', end_date='2026-07-15',
                           project_manager='张工', remarks=''),
            ConstructionPlan(project_name='XX污水处理厂扩建工程', unit_project='生化处理车间',
                           drawing_content='给排水管线图', major='给排水', designer='赵工',
                           major_director='赵总', estimated_quantitative='给排水专业全套施工图',
                           a1_drawing_count='80', plan_completion_percent='100%',
                           plan_work_days=76, completion_status='已完成', start_date='2026-02-15', end_date='2026-05-01',
                           project_manager='张工', remarks='已完成审查'),
            ConstructionPlan(project_name='YY热电厂改造项目', unit_project='主厂房',
                           drawing_content='暖通空调施工图', major='暖通', designer='陈工',
                           major_director='陈总', estimated_quantitative='暖通专业全套施工图',
                           a1_drawing_count='90', plan_completion_percent='45%',
                           plan_work_days=140, completion_status='进行中', start_date='2026-03-15', end_date='2026-08-01',
                           project_manager='刘工', remarks='含洁净室设计'),
            ConstructionPlan(project_name='YY热电厂改造项目', unit_project='主厂房',
                           drawing_content='电气施工图', major='电气', designer='孙工',
                           major_director='孙总', estimated_quantitative='电气专业全套施工图',
                           a1_drawing_count='100', plan_completion_percent='80%',
                           plan_work_days=106, completion_status='已延期', start_date='2026-03-01', end_date='2026-06-15',
                           project_manager='刘工', remarks='设备资料延迟'),
        ]
        db.session.add_all(samples)

    # Create sample data for phase plans
    if PhasePlan.query.count() == 0:
        samples = [
            PhasePlan(project_name='XX污水处理厂扩建工程', project_manager='张工', design_phase='初步设计',
                     product_drawing_name='初步设计文件', designer='张工',
                     major='建筑、结构、给排水', major_director='李所长',
                     estimated_quantitative='500页', a1_drawing_count='120',
                     manual_page_count='200页', plan_completion_percent='100%',
                     plan_work_days=88, plan_start_date='2026-02-01', plan_end_date='2026-04-30',
                     monthly_deviation='0', remarks='已通过评审'),
            PhasePlan(project_name='XX污水处理厂扩建工程', project_manager='张工', design_phase='施工图设计',
                     product_drawing_name='施工图设计文件', designer='张工',
                     major='建筑、结构、给排水、暖通、电气', major_director='李所长',
                     estimated_quantitative='800页', a1_drawing_count='200',
                     manual_page_count='300页', plan_completion_percent='60%',
                     plan_work_days=61, plan_start_date='2026-05-01', plan_end_date='2026-07-01',
                     monthly_deviation='+5', remarks='进行中'),
            PhasePlan(project_name='YY热电厂改造项目', project_manager='刘工', design_phase='方案设计',
                     product_drawing_name='方案设计文件', designer='刘工',
                     major='建筑、结构、电气', major_director='王所长',
                     estimated_quantitative='300页', a1_drawing_count='80',
                     manual_page_count='100页', plan_completion_percent='100%',
                     plan_work_days=60, plan_start_date='2026-01-15', plan_end_date='2026-03-15',
                     monthly_deviation='0', remarks='方案已确认'),
            PhasePlan(project_name='ZZ工业园区基础设施', project_manager='王工', design_phase='初步设计',
                     product_drawing_name='初步设计文件', designer='王工',
                     major='总图、建筑、结构、给排水', major_director='赵所长',
                     estimated_quantitative='400页', a1_drawing_count='100',
                     manual_page_count='150页', plan_completion_percent='30%',
                     plan_work_days=45, plan_start_date='2026-04-01', plan_end_date='2026-05-15',
                     monthly_deviation='-3', remarks='新项目'),
        ]
        db.session.add_all(samples)

    # Create sample data for technical plans
    if TechnicalPlan.query.count() == 0:
        samples = [
            TechnicalPlan(project_name='XX污水处理厂扩建工程', project_manager='张工', design_phase='施工图设计',
                         product_drawing_name='离心式鼓风机技术要求', designer='赵工',
                         major='电气', major_director='李所长',
                         estimated_quantitative='100页', a1_drawing_count='20',
                         manual_page_count='50页', plan_completion_percent='100%',
                         plan_work_days=30, plan_start_date='2026-03-01', plan_end_date='2026-03-30',
                         monthly_deviation='0', remarks='已完成'),
            TechnicalPlan(project_name='XX污水处理厂扩建工程', project_manager='张工', design_phase='施工图设计',
                         product_drawing_name='潜水排污泵技术要求', designer='赵工',
                         major='给排水', major_director='王所长',
                         estimated_quantitative='80页', a1_drawing_count='15',
                         manual_page_count='40页', plan_completion_percent='80%',
                         plan_work_days=25, plan_start_date='2026-03-05', plan_end_date='2026-03-30',
                         monthly_deviation='+2', remarks='进行中'),
            TechnicalPlan(project_name='YY热电厂改造项目', project_manager='刘工', design_phase='初步设计',
                         product_drawing_name='板式换热器技术要求', designer='陈工',
                         major='暖通', major_director='赵所长',
                         estimated_quantitative='60页', a1_drawing_count='10',
                         manual_page_count='30页', plan_completion_percent='50%',
                         plan_work_days=20, plan_start_date='2026-04-01', plan_end_date='2026-04-20',
                         monthly_deviation='-1', remarks='新项目'),
            TechnicalPlan(project_name='ZZ工业园区基础设施', project_manager='王工', design_phase='方案设计',
                         product_drawing_name='PLC控制柜技术要求', designer='孙工',
                         major='电气', major_director='李所长',
                         estimated_quantitative='120页', a1_drawing_count='25',
                         manual_page_count='60页', plan_completion_percent='100%',
                         plan_work_days=15, plan_start_date='2026-03-01', plan_end_date='2026-03-15',
                         monthly_deviation='0', remarks='已完成'),
        ]
        db.session.add_all(samples)

    db.session.commit()


# ==================== Page Routes ====================

@app.route('/')
def index():
    """返回主页面"""
    resp = make_response(send_from_directory('.', 'index.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


# ==================== Auth API ====================

@app.route('/api/login', methods=['POST'])
def api_login():
    """用户登录"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not username or not password:
        return jsonify({'code': 400, 'msg': '用户名和密码不能为空'}), 400

    user = User.query.filter_by(username=username, is_active=True).first()
    if not user or not user.check_password(password):
        return jsonify({'code': 401, 'msg': '用户名或密码错误'}), 401

    session['user_id'] = user.id
    session['username'] = user.username

    return jsonify({
        'code': 200,
        'msg': '登录成功',
        'data': user.to_dict()
    })


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """用户登出"""
    session.clear()
    return jsonify({'code': 200, 'msg': '已退出登录'})


@app.route('/api/current-user', methods=['GET'])
def api_current_user():
    """获取当前登录用户信息"""
    user = get_current_user()
    if not user:
        return jsonify({'code': 401, 'msg': '未登录'}), 401
    return jsonify({'code': 200, 'data': user.to_dict()})


# ==================== User Management API ====================

@app.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    """获取用户列表（仅管理员）"""
    user = get_current_user()
    if user.role != 'admin':
        return jsonify({'code': 403, 'msg': '仅管理员可查看用户列表'}), 403

    users = User.query.order_by(User.id).all()
    return jsonify({
        'code': 200,
        'data': [u.to_dict() for u in users]
    })


@app.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    """创建用户"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    real_name = data.get('realName', '').strip()
    role = data.get('role', 'staff')

    if not username or not password or not real_name:
        return jsonify({'code': 400, 'msg': '用户名、密码和姓名不能为空'}), 400

    if role not in ('admin', 'engineer', 'staff'):
        return jsonify({'code': 400, 'msg': '无效的角色类型'}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({'code': 400, 'msg': '用户名已存在'}), 400

    user = User(username=username, real_name=real_name, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    return jsonify({'code': 200, 'msg': '用户创建成功', 'data': user.to_dict()})


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def api_update_user(user_id):
    """更新用户"""
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'code': 404, 'msg': '用户不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    if 'realName' in data:
        user.real_name = data['realName'].strip()
    if 'role' in data and data['role'] in ('admin', 'engineer', 'staff'):
        user.role = data['role']
    if 'isActive' in data:
        user.is_active = data['isActive']
    if 'password' in data and data['password'].strip():
        user.set_password(data['password'].strip())

    db.session.commit()
    return jsonify({'code': 200, 'msg': '用户更新成功', 'data': user.to_dict()})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_delete_user(user_id):
    """删除用户"""
    if user_id == session.get('user_id'):
        return jsonify({'code': 400, 'msg': '不能删除当前登录的用户'}), 400

    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'code': 404, 'msg': '用户不存在'}), 404

    db.session.delete(user)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '用户已删除'})


@app.route('/api/users/sync-from-designers', methods=['POST'])
@admin_required
def api_sync_users_from_designers():
    """从设计人员同步用户（自动创建启用状态的设计人员为用户）"""
    # 获取所有启用状态的设计人员
    active_designers = Designer.query.filter_by(is_active=True).all()
    
    # 获取现有用户名
    existing_usernames = {u.username for u in User.query.all()}
    
    created_count = 0
    skipped_count = 0
    
    for designer in active_designers:
        # 生成用户名：从姓名拼音生成，如果没有拼音则用姓名
        if designer.name_pinyin:
            username = designer.name_pinyin
        else:
            username = generate_username(designer.name)
        
        # 如果用户名已存在，尝试添加数字后缀
        original_username = username
        suffix = 1
        while username in existing_usernames:
            username = f"{original_username}{suffix}"
            suffix += 1
        
        # 检查是否已存在该用户（根据real_name判断）
        if User.query.filter_by(real_name=designer.name).first():
            skipped_count += 1
            continue
        
        # 创建用户，密码为姓名拼音+123
        password = f"{username}123"
        user = User(
            username=username,
            real_name=designer.name,
            role='engineer'  # 设计人员身份为设计人员角色
        )
        user.set_password(password)
        db.session.add(user)
        existing_usernames.add(username)
        created_count += 1
    
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'msg': f'同步完成！新增 {created_count} 个用户，跳过 {skipped_count} 个已有用户'
    })


@app.route('/api/users/change-password', methods=['POST'])
@login_required
def api_change_password():
    """用户修改密码"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400
    
    old_password = data.get('oldPassword', '').strip()
    new_password = data.get('newPassword', '').strip()
    
    if not old_password or not new_password:
        return jsonify({'code': 400, 'msg': '请输入原密码和新密码'}), 400
    
    if len(new_password) < 6:
        return jsonify({'code': 400, 'msg': '新密码长度至少为6位'}), 400
    
    user = db.session.get(User, session.get('user_id'))
    if not user:
        return jsonify({'code': 404, 'msg': '用户不存在'}), 404
    
    if not user.check_password(old_password):
        return jsonify({'code': 400, 'msg': '原密码不正确'}), 400
    
    user.set_password(new_password)
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '密码修改成功'})


# ==================== Designer API ====================

@app.route('/api/designers', methods=['GET'])
@login_required
def api_get_designers():
    """获取所有设计人员列表（支持keyword搜索name）"""
    keyword = request.args.get('keyword', '').strip()
    query = Designer.query.order_by(Designer.id)
    if keyword:
        query = query.filter(Designer.name.contains(keyword))
    designers = query.all()
    return jsonify({'code': 200, 'data': [d.to_dict() for d in designers]})


@app.route('/api/designers', methods=['POST'])
@edit_required
def api_create_designer():
    """新增设计人员（admin/engineer）"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'code': 400, 'msg': '姓名不能为空'}), 400

    # 处理姓名拼音：如果未提供则自动生成
    name_pinyin = data.get('namePinyin', '').strip()
    if not name_pinyin:
        name_pinyin = generate_name_pinyin(name)

    designer = Designer(
        name=name,
        name_pinyin=name_pinyin,
        department=data.get('department', '').strip(),
        major=data.get('major', '').strip(),
        title=data.get('title', '').strip(),
        phone=data.get('phone', '').strip(),
    )
    db.session.add(designer)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '设计人员添加成功', 'data': designer.to_dict()})


@app.route('/api/designers/<int:designer_id>', methods=['PUT'])
@edit_required
def api_update_designer(designer_id):
    """更新设计人员（admin/engineer）"""
    designer = db.session.get(Designer, designer_id)
    if not designer:
        return jsonify({'code': 404, 'msg': '设计人员不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    if 'name' in data:
        designer.name = data['name'].strip()
    if 'namePinyin' in data:
        name_pinyin = data['namePinyin'].strip()
        if not name_pinyin:
            name_pinyin = generate_name_pinyin(designer.name)
        designer.name_pinyin = name_pinyin
    if 'department' in data:
        designer.department = data['department'].strip()
    if 'major' in data:
        designer.major = data['major'].strip()
    if 'title' in data:
        designer.title = data['title'].strip()
    if 'phone' in data:
        designer.phone = data['phone'].strip()
    if 'isActive' in data:
        designer.is_active = data['isActive']

    db.session.commit()
    return jsonify({'code': 200, 'msg': '设计人员更新成功', 'data': designer.to_dict()})


@app.route('/api/designers/<int:designer_id>', methods=['DELETE'])
@admin_required
def api_delete_designer(designer_id):
    """删除设计人员（admin）"""
    designer = db.session.get(Designer, designer_id)
    if not designer:
        return jsonify({'code': 404, 'msg': '设计人员不存在'}), 404

    db.session.delete(designer)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '设计人员已删除'})


@app.route('/api/designers/toggle-active/<int:designer_id>', methods=['POST'])
@edit_required
def api_toggle_designer_active(designer_id):
    """切换设计人员启用/禁用状态（admin/engineer）"""
    designer = db.session.get(Designer, designer_id)
    if not designer:
        return jsonify({'code': 404, 'msg': '设计人员不存在'}), 404

    designer.is_active = not designer.is_active
    db.session.commit()
    return jsonify({'code': 200, 'msg': '状态已更新', 'data': designer.to_dict()})


@app.route('/api/designers/export-template', methods=['GET'])
@login_required
def api_export_designer_template():
    """导出设计人员导入模板"""
    headers = ['姓名', '姓名拼音', '所属部门', '专业', '职称', '联系电话']
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('设计人员模板')
    
    # 设置表头样式
    header_style = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#f0f0f0',
        'border': 1
    })
    
    # 写入表头
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_style)
    
    # 设置列宽
    worksheet.set_column(0, 0, 15)  # 姓名
    worksheet.set_column(1, 1, 20)  # 姓名拼音
    worksheet.set_column(2, 2, 20)  # 所属部门
    worksheet.set_column(3, 3, 15)  # 专业
    worksheet.set_column(4, 4, 15)  # 职称
    worksheet.set_column(5, 5, 18)  # 联系电话
    
    workbook.close()
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='设计人员导入模板.xlsx',
        as_attachment=True
    )


@app.route('/api/designers/export-excel', methods=['GET'])
@login_required
def api_export_designers_excel():
    """导出设计人员数据到Excel"""
    designers = Designer.query.order_by(Designer.id).all()
    
    headers = ['序号', '姓名', '姓名拼音', '所属部门', '专业', '职称', '联系电话', '状态']
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('设计人员')
    
    # 设置表头样式
    header_style = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#f0f0f0',
        'border': 1
    })
    
    # 设置状态样式
    active_style = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'color': '#10b981', 'border': 1})
    inactive_style = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'color': '#6b7280', 'border': 1})
    
    # 写入表头
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_style)
    
    # 写入数据
    row_num = 1
    for designer in designers:
        worksheet.write(row_num, 0, row_num)
        worksheet.write(row_num, 1, designer.name)
        worksheet.write(row_num, 2, designer.name_pinyin or '')
        worksheet.write(row_num, 3, designer.department or '')
        worksheet.write(row_num, 4, designer.major or '')
        worksheet.write(row_num, 5, designer.title or '')
        worksheet.write(row_num, 6, designer.phone or '')
        worksheet.write(row_num, 7, '启用' if designer.is_active else '禁用', active_style if designer.is_active else inactive_style)
        row_num += 1
    
    # 设置列宽
    worksheet.set_column(0, 0, 8)   # 序号
    worksheet.set_column(1, 1, 12)  # 姓名
    worksheet.set_column(2, 2, 20)  # 姓名拼音
    worksheet.set_column(3, 3, 18)  # 所属部门
    worksheet.set_column(4, 4, 12)  # 专业
    worksheet.set_column(5, 5, 12)  # 职称
    worksheet.set_column(6, 6, 16)  # 联系电话
    worksheet.set_column(7, 7, 10)  # 状态
    
    workbook.close()
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='设计人员数据.xlsx',
        as_attachment=True
    )


@app.route('/api/designers/import-excel', methods=['POST'])
@edit_required
def api_import_designers_excel():
    """从Excel导入设计人员数据"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': '请选择要导入的文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'msg': '请选择要导入的文件'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'msg': '仅支持Excel文件（.xlsx/.xls）'}), 400
    
    try:
        # 读取Excel文件
        if file.filename.endswith('.xlsx'):
            workbook = load_workbook(file)
        else:
            workbook = xlrd.open_workbook(file_contents=file.read())
        
        worksheet = workbook.active if hasattr(workbook, 'active') else workbook.sheet_by_index(0)
        
        # 获取所有现有设计人员用于重复检测
        existing_designers = Designer.query.all()
        existing_keys = {}
        for d in existing_designers:
            key = f"{d.name}_{d.major}_{d.department}_{d.title}"
            existing_keys[key] = d
        
        imported_count = 0
        updated_count = 0
        
        # 从第二行开始读取（跳过表头）
        start_row = 1
        for row_idx in range(start_row, worksheet.max_row if hasattr(worksheet, 'max_row') else worksheet.nrows):
            try:
                # 获取单元格值
                # 列: 1-姓名, 2-姓名拼音, 3-所属部门, 4-专业, 5-职称, 6-联系电话
                if hasattr(worksheet, 'cell'):
                    name = str(worksheet.cell(row=row_idx+1, column=1).value).strip() if worksheet.cell(row=row_idx+1, column=1).value else ''
                    name_pinyin = str(worksheet.cell(row=row_idx+1, column=2).value).strip() if worksheet.cell(row=row_idx+1, column=2).value else ''
                    department = str(worksheet.cell(row=row_idx+1, column=3).value).strip() if worksheet.cell(row=row_idx+1, column=3).value else ''
                    major = str(worksheet.cell(row=row_idx+1, column=4).value).strip() if worksheet.cell(row=row_idx+1, column=4).value else ''
                    title = str(worksheet.cell(row=row_idx+1, column=5).value).strip() if worksheet.cell(row=row_idx+1, column=5).value else ''
                    phone = str(worksheet.cell(row=row_idx+1, column=6).value).strip() if worksheet.cell(row=row_idx+1, column=6).value else ''
                else:
                    name = str(worksheet.cell(row_idx, 0).value).strip() if worksheet.cell(row_idx, 0).value else ''
                    name_pinyin = str(worksheet.cell(row_idx, 1).value).strip() if worksheet.cell(row_idx, 1).value else ''
                    department = str(worksheet.cell(row_idx, 2).value).strip() if worksheet.cell(row_idx, 2).value else ''
                    major = str(worksheet.cell(row_idx, 3).value).strip() if worksheet.cell(row_idx, 3).value else ''
                    title = str(worksheet.cell(row_idx, 4).value).strip() if worksheet.cell(row_idx, 4).value else ''
                    phone = str(worksheet.cell(row_idx, 5).value).strip() if worksheet.cell(row_idx, 5).value else ''
                
                # 跳过空行
                if not name:
                    continue
                
                # 如果姓名拼音为空，自动生成
                if not name_pinyin:
                    name_pinyin = generate_name_pinyin(name)
                
                # 检查重复
                key = f"{name}_{major}_{department}_{title}"
                
                if key in existing_keys:
                    # 重复数据，覆盖更新
                    existing_designer = existing_keys[key]
                    existing_designer.name_pinyin = name_pinyin
                    existing_designer.department = department
                    existing_designer.major = major
                    existing_designer.title = title
                    existing_designer.phone = phone
                    existing_designer.is_active = True
                    updated_count += 1
                else:
                    # 新数据
                    new_designer = Designer(
                        name=name,
                        name_pinyin=name_pinyin,
                        department=department,
                        major=major,
                        title=title,
                        phone=phone,
                        is_active=True
                    )
                    db.session.add(new_designer)
                    imported_count += 1
            
            except Exception as e:
                continue
        
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'msg': f'导入完成！新增 {imported_count} 人，更新 {updated_count} 人'
        })
    
    except Exception as e:
        return jsonify({'code': 500, 'msg': f'导入失败：{str(e)}'}), 500


@app.route('/api/designers/suggestions', methods=['GET'])
@login_required
def api_designer_suggestions():
    """联想搜索（根据首字匹配name，返回前10条）"""
    keyword = request.args.get('keyword', '').strip()
    if not keyword:
        return jsonify({'code': 200, 'data': []})

    designers = Designer.query.filter(
        Designer.is_active == True,
        Designer.name.startswith(keyword)
    ).order_by(Designer.name).limit(10).all()

    return jsonify({'code': 200, 'data': [d.to_dict() for d in designers]})


# ==================== Major Category API ====================

@app.route('/api/major-categories', methods=['GET'])
@login_required
def api_get_major_categories():
    """获取专业类别列表"""
    categories = MajorCategory.query.order_by(MajorCategory.id).all()
    return jsonify({'code': 200, 'data': [c.to_dict() for c in categories]})


@app.route('/api/major-categories', methods=['POST'])
@edit_required
def api_create_major_category():
    """新增专业类别（admin/engineer）"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'code': 400, 'msg': '专业类别名称不能为空'}), 400

    if MajorCategory.query.filter_by(name=name).first():
        return jsonify({'code': 400, 'msg': '该专业类别已存在'}), 400

    category = MajorCategory(name=name)
    db.session.add(category)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '专业类别添加成功', 'data': category.to_dict()})


@app.route('/api/major-categories/<int:category_id>', methods=['DELETE'])
@edit_required
def api_delete_major_category(category_id):
    """删除专业类别（admin/engineer）"""
    category = db.session.get(MajorCategory, category_id)
    if not category:
        return jsonify({'code': 404, 'msg': '专业类别不存在'}), 404

    category_name = category.name
    db.session.delete(category)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '专业类别已删除'})


# ==================== Construction Plan API ====================

@app.route('/api/construction', methods=['GET'])
@login_required
def api_get_construction():
    """获取施工图计划列表"""
    keyword = request.args.get('keyword', '').strip()
    query = ConstructionPlan.query.order_by(ConstructionPlan.id)
    if keyword:
        query = query.filter(db.or_(
            ConstructionPlan.project_name.contains(keyword),
            ConstructionPlan.unit_project.contains(keyword),
            ConstructionPlan.drawing_content.contains(keyword),
            ConstructionPlan.major.contains(keyword),
            ConstructionPlan.designer.contains(keyword),
            ConstructionPlan.project_manager.contains(keyword),
        ))
    records = query.all()
    return jsonify({'code': 200, 'data': [r.to_dict() for r in records]})


@app.route('/api/construction', methods=['POST'])
@edit_required
def api_create_construction():
    """新增施工图计划"""
    data = request.get_json()
    if not data or not data.get('projectName'):
        return jsonify({'code': 400, 'msg': '项目名称不能为空'}), 400
    
    # 计算计划工作日
    plan_work_days = data.get('planWorkDays')
    if not plan_work_days and data.get('startDate') and data.get('endDate'):
        try:
            start = datetime.strptime(data.get('startDate'), '%Y-%m-%d')
            end = datetime.strptime(data.get('endDate'), '%Y-%m-%d')
            plan_work_days = (end - start).days
        except:
            plan_work_days = None
    
    record = ConstructionPlan(
        project_name=data.get('projectName', ''),
        project_manager=data.get('projectManager', ''),
        unit_project=data.get('unitProject', ''),
        drawing_content=data.get('drawingContent', ''),
        designer=data.get('designer', ''),
        major=data.get('major', ''),
        major_director=data.get('majorDirector', ''),
        estimated_quantitative=data.get('estimatedQuantitative', ''),
        a1_drawing_count=data.get('a1DrawingCount', ''),
        plan_completion_percent=data.get('planCompletionPercent', ''),
        plan_work_days=plan_work_days,
        start_date=data.get('startDate', ''),
        end_date=data.get('endDate', ''),
        monthly_deviation=data.get('monthlyDeviation', ''),
        completion_status=data.get('completionStatus', ''),
        remarks=data.get('remarks', ''),
        created_by=session.get('user_id')
    )
    db.session.add(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '添加成功', 'data': record.to_dict()})


@app.route('/api/construction/<int:record_id>', methods=['PUT'])
@edit_required
def api_update_construction(record_id):
    """更新施工图计划"""
    record = db.session.get(ConstructionPlan, record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    record.project_name = data.get('projectName', record.project_name)
    record.project_manager = data.get('projectManager', record.project_manager)
    record.unit_project = data.get('unitProject', record.unit_project)
    record.drawing_content = data.get('drawingContent', record.drawing_content)
    record.designer = data.get('designer', record.designer)
    record.major = data.get('major', record.major)
    record.major_director = data.get('majorDirector', record.major_director)
    record.estimated_quantitative = data.get('estimatedQuantitative', record.estimated_quantitative)
    record.a1_drawing_count = data.get('a1DrawingCount', record.a1_drawing_count)
    record.plan_completion_percent = data.get('planCompletionPercent', record.plan_completion_percent)
    record.start_date = data.get('startDate', record.start_date)
    record.end_date = data.get('endDate', record.end_date)
    record.monthly_deviation = data.get('monthlyDeviation', record.monthly_deviation)
    record.completion_status = data.get('completionStatus', record.completion_status)
    record.remarks = data.get('remarks', record.remarks)
    
    # 更新计划工作日
    plan_work_days = data.get('planWorkDays')
    if plan_work_days is not None:
        record.plan_work_days = plan_work_days
    elif data.get('startDate') and data.get('endDate'):
        try:
            start = datetime.strptime(data.get('startDate'), '%Y-%m-%d')
            end = datetime.strptime(data.get('endDate'), '%Y-%m-%d')
            record.plan_work_days = (end - start).days
        except:
            pass

    db.session.commit()
    return jsonify({'code': 200, 'msg': '更新成功', 'data': record.to_dict()})


@app.route('/api/construction/<int:record_id>', methods=['DELETE'])
@edit_required
def api_delete_construction(record_id):
    """删除施工图计划"""
    record = db.session.get(ConstructionPlan, record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '删除成功'})


# ==================== Phase Plan API ====================

@app.route('/api/phase', methods=['GET'])
@login_required
def api_get_phase():
    """获取阶段设计计划列表"""
    keyword = request.args.get('keyword', '').strip()
    query = PhasePlan.query.order_by(PhasePlan.id)
    if keyword:
        query = query.filter(db.or_(
            PhasePlan.project_name.contains(keyword),
            PhasePlan.design_phase.contains(keyword),
            PhasePlan.project_manager.contains(keyword),
            PhasePlan.product_drawing_name.contains(keyword),
            PhasePlan.designer.contains(keyword),
            PhasePlan.major.contains(keyword),
        ))
    records = query.all()
    return jsonify({'code': 200, 'data': [r.to_dict() for r in records]})


@app.route('/api/phase', methods=['POST'])
@edit_required
def api_create_phase():
    """新增阶段设计计划"""
    data = request.get_json()
    if not data or not data.get('projectName'):
        return jsonify({'code': 400, 'msg': '项目名称不能为空'}), 400

    # 计算计划工作日（如果未提供）
    plan_work_days = data.get('planWorkDays')
    plan_start_date = data.get('planStartDate', '')
    plan_end_date = data.get('planEndDate', '')
    if not plan_work_days and plan_start_date and plan_end_date:
        try:
            start = datetime.strptime(plan_start_date, '%Y-%m-%d')
            end = datetime.strptime(plan_end_date, '%Y-%m-%d')
            plan_work_days = (end - start).days
        except:
            plan_work_days = None
    elif plan_work_days:
        try:
            plan_work_days = int(plan_work_days)
        except:
            plan_work_days = None

    record = PhasePlan(
        project_name=data.get('projectName', ''),
        project_manager=data.get('projectManager', ''),
        design_phase=data.get('designPhase', ''),
        product_drawing_name=data.get('productDrawingName', ''),
        designer=data.get('designer', ''),
        major=data.get('major', ''),
        major_director=data.get('majorDirector', ''),
        estimated_quantitative=data.get('estimatedQuantitative', ''),
        a1_drawing_count=data.get('a1DrawingCount', ''),
        manual_page_count=data.get('manualPageCount', ''),
        plan_completion_percent=data.get('planCompletionPercent', ''),
        plan_work_days=plan_work_days,
        plan_start_date=plan_start_date,
        plan_end_date=plan_end_date,
        monthly_deviation=data.get('monthlyDeviation', ''),
        remarks=data.get('remarks', ''),
        created_by=session.get('user_id')
    )
    db.session.add(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '添加成功', 'data': record.to_dict()})


@app.route('/api/phase/<int:record_id>', methods=['PUT'])
@edit_required
def api_update_phase(record_id):
    """更新阶段设计计划"""
    record = db.session.get(PhasePlan, record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    record.project_name = data.get('projectName', record.project_name)
    record.project_manager = data.get('projectManager', record.project_manager)
    record.design_phase = data.get('designPhase', record.design_phase)
    record.product_drawing_name = data.get('productDrawingName', record.product_drawing_name)
    record.designer = data.get('designer', record.designer)
    record.major = data.get('major', record.major)
    record.major_director = data.get('majorDirector', record.major_director)
    record.estimated_quantitative = data.get('estimatedQuantitative', record.estimated_quantitative)
    record.a1_drawing_count = data.get('a1DrawingCount', record.a1_drawing_count)
    record.manual_page_count = data.get('manualPageCount', record.manual_page_count)
    record.plan_completion_percent = data.get('planCompletionPercent', record.plan_completion_percent)

    # 处理计划工作日
    plan_work_days = data.get('planWorkDays')
    plan_start_date = data.get('planStartDate', record.plan_start_date)
    plan_end_date = data.get('planEndDate', record.plan_end_date)
    if plan_start_date:
        record.plan_start_date = plan_start_date
    if plan_end_date:
        record.plan_end_date = plan_end_date
    if plan_work_days:
        try:
            record.plan_work_days = int(plan_work_days)
        except:
            record.plan_work_days = None
    elif plan_start_date and plan_end_date:
        try:
            start = datetime.strptime(plan_start_date, '%Y-%m-%d')
            end = datetime.strptime(plan_end_date, '%Y-%m-%d')
            record.plan_work_days = (end - start).days
        except:
            pass

    record.monthly_deviation = data.get('monthlyDeviation', record.monthly_deviation)
    record.remarks = data.get('remarks', record.remarks)

    db.session.commit()
    return jsonify({'code': 200, 'msg': '更新成功', 'data': record.to_dict()})


@app.route('/api/phase/<int:record_id>', methods=['DELETE'])
@edit_required
def api_delete_phase(record_id):
    """删除阶段设计计划"""
    record = db.session.get(PhasePlan, record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '删除成功'})


# ==================== Technical Plan API ====================

@app.route('/api/technical', methods=['GET'])
@login_required
def api_get_technical():
    """获取技术要求计划列表"""
    keyword = request.args.get('keyword', '').strip()
    project = request.args.get('project', '').strip()
    query = TechnicalPlan.query.order_by(TechnicalPlan.id)
    if keyword:
        query = query.filter(db.or_(
            TechnicalPlan.project_name.contains(keyword),
            TechnicalPlan.product_drawing_name.contains(keyword),
            TechnicalPlan.designer.contains(keyword),
            TechnicalPlan.major.contains(keyword),
        ))
    if project:
        query = query.filter(TechnicalPlan.project_name == project)
    records = query.all()
    return jsonify({'code': 200, 'data': [r.to_dict() for r in records]})


@app.route('/api/technical', methods=['POST'])
@edit_required
def api_create_technical():
    """新增技术要求计划"""
    data = request.get_json()
    if not data or not data.get('productDrawingName'):
        return jsonify({'code': 400, 'msg': '成品套图/文件名称不能为空'}), 400

    # 计算计划工作日（如果未提供）
    plan_work_days = data.get('planWorkDays')
    plan_start_date = data.get('planStartDate', '')
    plan_end_date = data.get('planEndDate', '')
    if not plan_work_days and plan_start_date and plan_end_date:
        try:
            start = datetime.strptime(plan_start_date, '%Y-%m-%d')
            end = datetime.strptime(plan_end_date, '%Y-%m-%d')
            plan_work_days = (end - start).days
        except:
            plan_work_days = None
    elif plan_work_days:
        try:
            plan_work_days = int(plan_work_days)
        except:
            plan_work_days = None

    record = TechnicalPlan(
        project_name=data.get('projectName', ''),
        project_manager=data.get('projectManager', ''),
        design_phase=data.get('designPhase', ''),
        product_drawing_name=data.get('productDrawingName', ''),
        designer=data.get('designer', ''),
        major=data.get('major', ''),
        major_director=data.get('majorDirector', ''),
        estimated_quantitative=data.get('estimatedQuantitative', ''),
        a1_drawing_count=data.get('a1DrawingCount', ''),
        manual_page_count=data.get('manualPageCount', ''),
        plan_completion_percent=data.get('planCompletionPercent', ''),
        plan_work_days=plan_work_days,
        plan_start_date=plan_start_date,
        plan_end_date=plan_end_date,
        monthly_deviation=data.get('monthlyDeviation', ''),
        remarks=data.get('remarks', ''),
        created_by=session.get('user_id')
    )
    db.session.add(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '添加成功', 'data': record.to_dict()})


@app.route('/api/technical/<int:record_id>', methods=['PUT'])
@edit_required
def api_update_technical(record_id):
    """更新技术要求计划"""
    record = db.session.get(TechnicalPlan, record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    record.project_name = data.get('projectName', record.project_name)
    record.project_manager = data.get('projectManager', record.project_manager)
    record.design_phase = data.get('designPhase', record.design_phase)
    record.product_drawing_name = data.get('productDrawingName', record.product_drawing_name)
    record.designer = data.get('designer', record.designer)
    record.major = data.get('major', record.major)
    record.major_director = data.get('majorDirector', record.major_director)
    record.estimated_quantitative = data.get('estimatedQuantitative', record.estimated_quantitative)
    record.a1_drawing_count = data.get('a1DrawingCount', record.a1_drawing_count)
    record.manual_page_count = data.get('manualPageCount', record.manual_page_count)
    record.plan_completion_percent = data.get('planCompletionPercent', record.plan_completion_percent)

    # 处理计划工作日
    plan_work_days = data.get('planWorkDays')
    plan_start_date = data.get('planStartDate', record.plan_start_date)
    plan_end_date = data.get('planEndDate', record.plan_end_date)
    if plan_start_date:
        record.plan_start_date = plan_start_date
    if plan_end_date:
        record.plan_end_date = plan_end_date
    if plan_work_days:
        try:
            record.plan_work_days = int(plan_work_days)
        except:
            record.plan_work_days = None
    elif plan_start_date and plan_end_date:
        try:
            start = datetime.strptime(plan_start_date, '%Y-%m-%d')
            end = datetime.strptime(plan_end_date, '%Y-%m-%d')
            record.plan_work_days = (end - start).days
        except:
            pass

    record.monthly_deviation = data.get('monthlyDeviation', record.monthly_deviation)
    record.remarks = data.get('remarks', record.remarks)

    db.session.commit()
    return jsonify({'code': 200, 'msg': '更新成功', 'data': record.to_dict()})


@app.route('/api/technical/<int:record_id>', methods=['DELETE'])
@edit_required
def api_delete_technical(record_id):
    """删除技术要求计划"""
    record = db.session.get(TechnicalPlan, record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '删除成功'})


# ==================== Batch Delete API ====================

@app.route('/api/batch-delete/construction', methods=['POST'])
@edit_required
def api_batch_delete_construction():
    """批量删除施工图计划"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    query = ConstructionPlan.query

    if 'project' in data and data['project']:
        query = query.filter(ConstructionPlan.project_name == data['project'])
    if 'unit_project' in data and data['unit_project']:
        query = query.filter(ConstructionPlan.unit_project == data['unit_project'])
    if 'designer' in data and data['designer']:
        query = query.filter(ConstructionPlan.designer == data['designer'])
    if 'major' in data and data['major']:
        query = query.filter(ConstructionPlan.major == data['major'])
    if 'start_date' in data and data['start_date']:
        query = query.filter(ConstructionPlan.end_date >= data['start_date'])
    if 'end_date' in data and data['end_date']:
        query = query.filter(ConstructionPlan.end_date <= data['end_date'])

    records = query.all()
    count = len(records)
    
    for record in records:
        db.session.delete(record)
    db.session.commit()

    return jsonify({'code': 200, 'msg': '批量删除成功', 'data': {'count': count}})


@app.route('/api/batch-delete/phase', methods=['POST'])
@edit_required
def api_batch_delete_phase():
    """批量删除阶段设计计划"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    query = PhasePlan.query

    if 'project' in data and data['project']:
        query = query.filter(PhasePlan.project_name == data['project'])
    if 'major' in data and data['major']:
        query = query.filter(PhasePlan.major == data['major'])
    if 'start_date' in data and data['start_date']:
        query = query.filter(PhasePlan.plan_end_date >= data['start_date'])
    if 'end_date' in data and data['end_date']:
        query = query.filter(PhasePlan.plan_end_date <= data['end_date'])

    records = query.all()
    count = len(records)
    
    for record in records:
        db.session.delete(record)
    db.session.commit()

    return jsonify({'code': 200, 'msg': '批量删除成功', 'data': {'count': count}})


@app.route('/api/batch-delete/technical', methods=['POST'])
@edit_required
def api_batch_delete_technical():
    """批量删除技术要求计划"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    query = TechnicalPlan.query

    if 'project' in data and data['project']:
        query = query.filter(TechnicalPlan.project_name == data['project'])
    if 'equipment' in data and data['equipment']:
        query = query.filter(TechnicalPlan.product_drawing_name.contains(data['equipment']))
    if 'designer' in data and data['designer']:
        query = query.filter(TechnicalPlan.designer == data['designer'])
    if 'start_date' in data and data['start_date']:
        query = query.filter(TechnicalPlan.plan_end_date >= data['start_date'])
    if 'end_date' in data and data['end_date']:
        query = query.filter(TechnicalPlan.plan_end_date <= data['end_date'])

    records = query.all()
    count = len(records)
    
    for record in records:
        db.session.delete(record)
    db.session.commit()

    return jsonify({'code': 200, 'msg': '批量删除成功', 'data': {'count': count}})


# ==================== Excel Template / Import / Export API ====================

# 表头映射定义
EXCEL_HEADERS = {
    'construction': [
        '项目名称', '所属单位', '项目负责人', '单位工程名称', '成品套图',
        '设计人', '专业', '专业所长', '预估量化指标', '折合A1图纸张数',
        '计划完成百分比', '计划工作日', '计划开始时间', '计划完成时间',
        '上月滚动偏差', '完成情况', '备注', '设计阶段'
    ],
    'phase': [
        '项目名称', '项目负责人', '设计阶段', '成品套图/文件名称', '设计人', '专业',
        '专业所长', '预估量化指标', '折合A1图纸张数', '说明书页数', '计划完成百分比',
        '计划工作日', '计划开始时间', '计划完成时间', '上月滚动偏差', '备注'
    ],
    'technical': [
        '项目名称', '项目负责人', '设计阶段', '成品套图/文件名称', '设计人',
        '专业', '专业所长', '预估量化指标', '折合A1图纸张数', '说明书页数',
        '计划完成百分比', '计划工作日', '计划开始时间', '计划完成时间',
        '上月滚动偏差', '备注'
    ],
    'workload': [
        '月份', '所属单位', '姓名', '专业', '专业所长', '人员角色',
        '具体完成工作内容', '量化指标', '项目名称', '项目负责人',
        '设计阶段', '成品套图/文件名称', '折合A1图纸张数', '说明书页数',
        '套图/文件完成率', '实际工作日', '工作开始时间', '工作结束时间',
        '计划偏差(天)', '备注'
    ],
}

EXCEL_FILENAMES = {
    'construction': '施工图计划模板.xlsx',
    'phase': '阶段设计计划模板.xlsx',
    'technical': '技术要求计划模板.xlsx',
}


@app.route('/api/template/<table_type>', methods=['GET'])
@login_required
def api_download_template(table_type):
    """下载Excel模板"""
    if table_type not in EXCEL_HEADERS:
        return jsonify({'code': 400, 'msg': '无效的表格类型，支持: construction / phase / technical'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = table_type
    headers = EXCEL_HEADERS[table_type]
    ws.append(headers)

    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 自动调整列宽
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(header) * 2, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = EXCEL_FILENAMES.get(table_type, f'{table_type}_template.xlsx')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/import/<table_type>', methods=['POST'])
@edit_required
def api_import_excel(table_type):
    """导入Excel数据"""
    if table_type not in EXCEL_HEADERS:
        return jsonify({'code': 400, 'msg': '无效的表格类型，支持: construction / phase / technical'}), 400

    data = request.form
    overwrite = data.get('overwrite', 'false').lower() == 'true'

    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': '请上传文件（字段名: file）'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 400, 'msg': '文件名为空'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'code': 400, 'msg': '仅支持 .xlsx 格式的Excel文件'}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'code': 400, 'msg': f'Excel文件解析失败: {str(e)}'}), 400

    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    if len(rows) < 2:
        return jsonify({'code': 400, 'msg': 'Excel文件没有数据行（至少需要表头+1行数据）'}), 400

    # 验证表头
    header_row = [str(cell).strip() if cell else '' for cell in rows[0]]
    expected_headers = EXCEL_HEADERS[table_type]
    if header_row != expected_headers:
        return jsonify({
            'code': 400,
            'msg': f'表头不匹配。期望: {expected_headers}，实际: {header_row}'
        }), 400

    success_count = 0
    updated_count = 0
    fail_count = 0
    fail_rows = []
    duplicates = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            values = [str(cell).strip() if cell is not None else '' for cell in row]

            if table_type == 'construction':
                if not values[0]:
                    fail_count += 1
                    fail_rows.append(row_idx)
                    continue
                
                # 检查是否存在重复记录（根据项目名称、单位工程、成品套图、设计人）
                existing = ConstructionPlan.query.filter(
                    ConstructionPlan.project_name == values[0],
                    ConstructionPlan.unit_project == values[3],
                    ConstructionPlan.drawing_content == values[4],
                    ConstructionPlan.designer == values[5]
                ).first()
                
                def parse_date(date_str):
                    """尝试多种日期格式解析"""
                    formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']
                    for fmt in formats:
                        try:
                            return datetime.strptime(date_str, fmt)
                        except ValueError:
                            continue
                    return None
                
                plan_work_days = values[11]
                if not plan_work_days and values[12] and values[13]:
                    start_date = parse_date(values[12])
                    end_date = parse_date(values[13])
                    if start_date and end_date:
                        plan_work_days = (end_date - start_date).days
                    else:
                        plan_work_days = None
                elif plan_work_days:
                    try:
                        plan_work_days = int(plan_work_days)
                    except:
                        plan_work_days = None
                
                if existing:
                    if overwrite:
                        # 覆盖更新
                        existing.project_manager = values[2]
                        existing.unit_project = values[3]
                        existing.drawing_content = values[4]
                        existing.designer = values[5]
                        existing.major = values[6]
                        existing.major_director = values[7]
                        existing.estimated_quantitative = values[8]
                        existing.a1_drawing_count = values[9]
                        existing.plan_completion_percent = values[10]
                        existing.plan_work_days = plan_work_days
                        existing.start_date = values[12]
                        existing.end_date = values[13]
                        existing.monthly_deviation = values[14]
                        existing.completion_status = values[15]
                        existing.remarks = values[16]
                        existing.updated_at = datetime.now()
                        updated_count += 1
                    else:
                        duplicates.append({
                            'row': row_idx,
                            'project': values[0],
                            'unit': values[3],
                            'drawing': values[4],
                            'designer': values[5]
                        })
                        continue
                else:
                    record = ConstructionPlan(
                        project_name=values[0],
                        project_manager=values[2],
                        unit_project=values[3],
                        drawing_content=values[4],
                        designer=values[5],
                        major=values[6],
                        major_director=values[7],
                        estimated_quantitative=values[8],
                        a1_drawing_count=values[9],
                        plan_completion_percent=values[10],
                        plan_work_days=plan_work_days,
                        start_date=values[12],
                        end_date=values[13],
                        monthly_deviation=values[14],
                        completion_status=values[15],
                        remarks=values[16],
                        created_by=session.get('user_id')
                    )
                    db.session.add(record)
                    success_count += 1
                    
            elif table_type == 'phase':
                if not values[0]:
                    fail_count += 1
                    fail_rows.append(row_idx)
                    continue
                # 计算计划工作日（如果未提供）
                plan_work_days = values[11]
                if not plan_work_days and values[12] and values[13]:
                    start_date = parse_date(values[12])
                    end_date = parse_date(values[13])
                    if start_date and end_date:
                        plan_work_days = (end_date - start_date).days
                    else:
                        plan_work_days = None
                elif plan_work_days:
                    try:
                        plan_work_days = int(plan_work_days)
                    except:
                        plan_work_days = None
                record = PhasePlan(
                    project_name=values[0],
                    project_manager=values[1],
                    design_phase=values[2],
                    product_drawing_name=values[3],
                    designer=values[4],
                    major=values[5],
                    major_director=values[6],
                    estimated_quantitative=values[7],
                    a1_drawing_count=values[8],
                    manual_page_count=values[9],
                    plan_completion_percent=values[10],
                    plan_work_days=plan_work_days,
                    plan_start_date=values[12],
                    plan_end_date=values[13],
                    monthly_deviation=values[14],
                    remarks=values[15],
                    created_by=session.get('user_id')
                )
                db.session.add(record)
                success_count += 1
                
            elif table_type == 'technical':
                if not values[3]:
                    fail_count += 1
                    fail_rows.append(row_idx)
                    continue
                # 计算计划工作日（如果未提供）
                plan_work_days = values[11]
                if not plan_work_days and values[12] and values[13]:
                    start_date = parse_date(values[12])
                    end_date = parse_date(values[13])
                    if start_date and end_date:
                        plan_work_days = (end_date - start_date).days
                    else:
                        plan_work_days = None
                elif plan_work_days:
                    try:
                        plan_work_days = int(plan_work_days)
                    except:
                        plan_work_days = None
                record = TechnicalPlan(
                    project_name=values[0],
                    project_manager=values[1],
                    design_phase=values[2],
                    product_drawing_name=values[3],
                    designer=values[4],
                    major=values[5],
                    major_director=values[6],
                    estimated_quantitative=values[7],
                    a1_drawing_count=values[8],
                    manual_page_count=values[9],
                    plan_completion_percent=values[10],
                    plan_work_days=plan_work_days,
                    plan_start_date=values[12],
                    plan_end_date=values[13],
                    monthly_deviation=values[14],
                    remarks=values[15],
                    created_by=session.get('user_id')
                )
                db.session.add(record)
                success_count += 1
                
            else:
                fail_count += 1
                fail_rows.append(row_idx)
                continue

        except Exception:
            fail_count += 1
            fail_rows.append(row_idx)

    # 如果发现重复且未选择覆盖，则返回需要确认
    if duplicates and not overwrite:
        return jsonify({
            'code': 300,
            'msg': f'发现 {len(duplicates)} 条重复记录',
            'data': {
                'duplicates': duplicates,
                'success': success_count,
                'fail': fail_count,
                'failRows': fail_rows
            }
        })

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'msg': f'数据库提交失败: {str(e)}'}), 500

    result_msg = f'导入完成：新增 {success_count} 条'
    if updated_count > 0:
        result_msg += f'，更新 {updated_count} 条'
    if fail_count > 0:
        result_msg += f'，失败 {fail_count} 条（行号: {fail_rows}）'

    return jsonify({
        'code': 200,
        'msg': result_msg,
        'data': {
            'success': success_count,
            'updated': updated_count,
            'fail': fail_count,
            'failRows': fail_rows
        }
    })


@app.route('/api/export-excel/<table_type>', methods=['GET'])
@login_required
def api_export_excel(table_type):
    """导出Excel"""
    if table_type not in EXCEL_HEADERS:
        return jsonify({'code': 400, 'msg': '无效的表格类型，支持: construction / phase / technical'}), 400

    keyword = request.args.get('keyword', '').strip()
    project = request.args.get('project', '').strip()
    major = request.args.get('major', '').strip()

    # 根据类型查询数据
    if table_type == 'construction':
        query = ConstructionPlan.query.order_by(ConstructionPlan.id)
        if keyword:
            query = query.filter(db.or_(
                ConstructionPlan.project_name.contains(keyword),
                ConstructionPlan.unit_project.contains(keyword),
                ConstructionPlan.drawing_content.contains(keyword),
                ConstructionPlan.major.contains(keyword),
                ConstructionPlan.designer.contains(keyword),
                ConstructionPlan.project_manager.contains(keyword),
            ))
        if project:
            query = query.filter(ConstructionPlan.project_name == project)
        if major:
            query = query.filter(ConstructionPlan.major == major)
        records = query.all()
        headers = EXCEL_HEADERS['construction']
        # 导出列顺序：项目名称、所属单位、项目负责人、单位工程名称、成品套图、设计人、专业、专业所长、预估量化指标、折合A1图纸张数、计划完成百分比、计划工作日、计划开始时间、计划完成时间、上月滚动偏差、完成情况、备注、设计阶段
        data_rows = [[
            r.project_name, '', r.project_manager, r.unit_project, r.drawing_content,
            r.designer, r.major, r.major_director, r.estimated_quantitative, r.a1_drawing_count,
            r.plan_completion_percent, r.plan_work_days if r.plan_work_days else '', r.start_date, r.end_date,
            r.monthly_deviation, r.completion_status, r.remarks, '施工图'
        ] for r in records]

    elif table_type == 'phase':
        query = PhasePlan.query.order_by(PhasePlan.id)
        if keyword:
            query = query.filter(db.or_(
                PhasePlan.project_name.contains(keyword),
                PhasePlan.design_phase.contains(keyword),
                PhasePlan.project_manager.contains(keyword),
                PhasePlan.product_drawing_name.contains(keyword),
                PhasePlan.designer.contains(keyword),
                PhasePlan.major.contains(keyword),
            ))
        if project:
            query = query.filter(PhasePlan.project_name == project)
        records = query.all()
        headers = EXCEL_HEADERS['phase']
        data_rows = [[
            r.project_name, r.project_manager, r.design_phase, r.product_drawing_name,
            r.designer, r.major, r.major_director, r.estimated_quantitative, r.a1_drawing_count,
            r.manual_page_count, r.plan_completion_percent,
            r.plan_work_days if r.plan_work_days else '',
            r.plan_start_date, r.plan_end_date, r.monthly_deviation, r.remarks
        ] for r in records]

    elif table_type == 'technical':
        query = TechnicalPlan.query.order_by(TechnicalPlan.id)
        if keyword:
            query = query.filter(db.or_(
                TechnicalPlan.project_name.contains(keyword),
                TechnicalPlan.product_drawing_name.contains(keyword),
                TechnicalPlan.designer.contains(keyword),
                TechnicalPlan.major.contains(keyword),
            ))
        if project:
            query = query.filter(TechnicalPlan.project_name == project)
        records = query.all()
        headers = EXCEL_HEADERS['technical']
        data_rows = [[
            r.project_name, r.project_manager, r.design_phase, r.product_drawing_name,
            r.designer, r.major, r.major_director, r.estimated_quantitative,
            r.a1_drawing_count, r.manual_page_count, r.plan_completion_percent,
            r.plan_work_days if r.plan_work_days else '',
            r.plan_start_date, r.plan_end_date, r.monthly_deviation, r.remarks
        ] for r in records]
    else:
        data_rows = []
        headers = []

    wb = Workbook()
    ws = wb.active
    ws.title = table_type

    # 写入表头
    ws.append(headers)
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 写入数据
    for row_data in data_rows:
        ws.append(row_data)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(str(headers[col_idx - 1])) * 2
        for row in data_rows:
            if col_idx <= len(row) and row[col_idx - 1]:
                max_len = max(max_len, len(str(row[col_idx - 1])) * 1.5)
        ws.column_dimensions[col_letter].width = min(max(max_len, 12), 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    export_filenames = {
        'construction': f'施工图计划_{timestamp}.xlsx',
        'phase': f'阶段设计计划_{timestamp}.xlsx',
        'technical': f'技术要求计划_{timestamp}.xlsx',
    }
    filename = export_filenames.get(table_type, f'{table_type}_{timestamp}.xlsx')

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== Statistics API ====================

@app.route('/api/stats/overview', methods=['GET'])
@login_required
def api_stats_overview():
    """总览统计"""
    keyword = request.args.get('keyword', '').strip()

    # 施工图统计
    cq = ConstructionPlan.query
    if keyword:
        cq = cq.filter(db.or_(
            ConstructionPlan.project_name.contains(keyword),
            ConstructionPlan.unit_project.contains(keyword),
            ConstructionPlan.designer.contains(keyword),
        ))
    construction_total = cq.count()
    construction_completed = cq.filter(ConstructionPlan.completion_status == '已完成').count()
    construction_ongoing = cq.filter(ConstructionPlan.completion_status == '进行中').count()
    construction_not_started = cq.filter(ConstructionPlan.completion_status == '未开始').count()
    construction_delayed = cq.filter(ConstructionPlan.completion_status == '已延期').count()

    # 阶段设计统计
    pq = PhasePlan.query
    if keyword:
        pq = pq.filter(db.or_(
            PhasePlan.project_name.contains(keyword),
            PhasePlan.design_phase.contains(keyword),
        ))
    phase_total = pq.count()

    # 技术要求统计
    tq = TechnicalPlan.query
    if keyword:
        tq = tq.filter(db.or_(
            TechnicalPlan.equipment_name.contains(keyword),
            TechnicalPlan.designer.contains(keyword),
        ))
    technical_total = tq.count()

    # 专业分布（从施工图表统计）
    major_stats = {}
    cq_all = ConstructionPlan.query
    if keyword:
        cq_all = cq_all.filter(db.or_(
            ConstructionPlan.project_name.contains(keyword),
            ConstructionPlan.unit_project.contains(keyword),
            ConstructionPlan.designer.contains(keyword),
        ))
    for r in cq_all.all():
        major = r.major or '未分类'
        if major not in major_stats:
            major_stats[major] = 0
        major_stats[major] += 1

    return jsonify({
        'code': 200,
        'data': {
            'tables': {
                'construction': construction_total,
                'phase': phase_total,
                'technical': technical_total,
            },
            'completionStatus': {
                'labels': ['已完成', '进行中', '未开始', '已延期'],
                'values': [construction_completed, construction_ongoing, construction_not_started, construction_delayed]
            },
            'majorDistribution': {
                'labels': list(major_stats.keys()),
                'values': list(major_stats.values())
            }
        }
    })


@app.route('/api/stats/by-project', methods=['GET'])
@login_required
def api_stats_by_project():
    """按项目统计（支持过滤）"""
    project_param = request.args.get('project', '').strip()
    unit_project_param = request.args.get('unit_project', '').strip()
    major_param = request.args.get('major', '').strip()

    cq = ConstructionPlan.query

    # 按项目过滤（支持逗号分隔多个）
    if project_param:
        project_list = [p.strip() for p in project_param.split(',') if p.strip()]
        if project_list:
            cq = cq.filter(ConstructionPlan.project_name.in_(project_list))

    # 按单位工程过滤
    if unit_project_param:
        cq = cq.filter(ConstructionPlan.unit_project == unit_project_param)

    # 按专业过滤
    if major_param:
        major_list = [m.strip() for m in major_param.split(',') if m.strip()]
        if major_list:
            cq = cq.filter(ConstructionPlan.major.in_(major_list))

    records = cq.all()

    project_stats = {}
    for r in records:
        pname = r.project_name or '未命名项目'
        if pname not in project_stats:
            project_stats[pname] = {'total': 0, 'completed': 0, 'ongoing': 0, 'notStarted': 0, 'delayed': 0, 'items': []}
        project_stats[pname]['total'] += 1
        status = r.completion_status or ''
        if status == '已完成':
            project_stats[pname]['completed'] += 1
        elif status == '进行中':
            project_stats[pname]['ongoing'] += 1
        elif status == '未开始':
            project_stats[pname]['notStarted'] += 1
        elif status == '已延期':
            project_stats[pname]['delayed'] += 1
        project_stats[pname]['items'].append({
            'unitProject': r.unit_project or '',
            'drawingContent': r.drawing_content or '',
            'major': r.major or '',
            'designer': r.designer or '',
            'completionStatus': r.completion_status or '',
        })

    labels = list(project_stats.keys())
    details = []
    for label in labels:
        s = project_stats[label]
        details.append({
            'projectName': label,
            'total': s['total'],
            'completed': s['completed'],
            'ongoing': s['ongoing'],
            'notStarted': s['notStarted'],
            'delayed': s['delayed'],
            'items': s['items'],
        })

    return jsonify({
        'code': 200,
        'data': {
            'labels': labels,
            'totals': [project_stats[l]['total'] for l in labels],
            'completed': [project_stats[l]['completed'] for l in labels],
            'details': details,
        }
    })


@app.route('/api/stats/by-unit', methods=['GET'])
@login_required
def api_stats_by_unit():
    """按单位工程统计"""
    keyword = request.args.get('keyword', '').strip()

    cq = ConstructionPlan.query
    if keyword:
        cq = cq.filter(db.or_(
            ConstructionPlan.project_name.contains(keyword),
            ConstructionPlan.unit_project.contains(keyword),
            ConstructionPlan.designer.contains(keyword),
        ))
    records = cq.all()

    unit_stats = {}
    for r in records:
        uname = r.unit_project or '未分类'
        if uname not in unit_stats:
            unit_stats[uname] = 0
        unit_stats[uname] += 1

    # 按数量降序排列
    sorted_units = sorted(unit_stats.items(), key=lambda x: x[1], reverse=True)

    return jsonify({
        'code': 200,
        'data': {
            'labels': [u[0] for u in sorted_units],
            'values': [u[1] for u in sorted_units]
        }
    })


@app.route('/api/stats/by-major', methods=['GET'])
@login_required
def api_stats_by_major():
    """按专业统计（支持过滤）"""
    project = request.args.get('project', '').strip()
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str = request.args.get('end_date', '').strip()
    major_param = request.args.get('major', '').strip()

    cq = ConstructionPlan.query

    # 按项目过滤
    if project:
        cq = cq.filter(ConstructionPlan.project_name == project)

    # 按日期范围过滤（基于endDate）
    if start_date_str:
        cq = cq.filter(ConstructionPlan.end_date >= start_date_str)
    if end_date_str:
        cq = cq.filter(ConstructionPlan.end_date <= end_date_str)

    # 按专业过滤
    if major_param:
        major_list = [m.strip() for m in major_param.split(',') if m.strip()]
        if major_list:
            cq = cq.filter(ConstructionPlan.major.in_(major_list))

    records = cq.all()

    major_stats = {}
    for r in records:
        major = r.major or '未分类'
        if major not in major_stats:
            major_stats[major] = {'total': 0, 'completed': 0, 'ongoing': 0, 'notStarted': 0, 'delayed': 0, 'items': []}
        major_stats[major]['total'] += 1
        status = r.completion_status or ''
        if status == '已完成':
            major_stats[major]['completed'] += 1
        elif status == '进行中':
            major_stats[major]['ongoing'] += 1
        elif status == '未开始':
            major_stats[major]['notStarted'] += 1
        elif status == '已延期':
            major_stats[major]['delayed'] += 1
        major_stats[major]['items'].append({
            'projectName': r.project_name or '',
            'unitProject': r.unit_project or '',
            'drawingContent': r.drawing_content or '',
            'designer': r.designer or '',
            'completionStatus': r.completion_status or '',
            'startDate': r.start_date or '',
            'endDate': r.end_date or '',
        })

    labels = list(major_stats.keys())
    details = []
    for label in labels:
        s = major_stats[label]
        details.append({
            'major': label,
            'total': s['total'],
            'completed': s['completed'],
            'ongoing': s['ongoing'],
            'notStarted': s['notStarted'],
            'delayed': s['delayed'],
            'items': s['items'],
        })

    return jsonify({
        'code': 200,
        'data': {
            'labels': labels,
            'totals': [major_stats[l]['total'] for l in labels],
            'completed': [major_stats[l]['completed'] for l in labels],
            'ongoing': [major_stats[l]['ongoing'] for l in labels],
            'notStarted': [major_stats[l]['notStarted'] for l in labels],
            'delayed': [major_stats[l]['delayed'] for l in labels],
            'details': details,
        }
    })


@app.route('/api/stats/by-designer', methods=['GET'])
@login_required
def api_stats_by_designer():
    """按设计人员统计（支持过滤）"""
    designer_param = request.args.get('designer', '').strip()
    project = request.args.get('project', '').strip()
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str = request.args.get('end_date', '').strip()

    # 施工图查询
    cq = ConstructionPlan.query
    if project:
        cq = cq.filter(ConstructionPlan.project_name == project)
    if start_date_str:
        cq = cq.filter(ConstructionPlan.end_date >= start_date_str)
    if end_date_str:
        cq = cq.filter(ConstructionPlan.end_date <= end_date_str)
    if designer_param:
        designer_list = [d.strip() for d in designer_param.split(',') if d.strip()]
        if designer_list:
            cq = cq.filter(ConstructionPlan.designer.in_(designer_list))
    construction_records = cq.all()

    # 技术要求查询
    tq = TechnicalPlan.query
    if designer_param:
        designer_list = [d.strip() for d in designer_param.split(',') if d.strip()]
        if designer_list:
            tq = tq.filter(TechnicalPlan.designer.in_(designer_list))
    technical_records = tq.all()

    designer_stats = {}
    for r in construction_records:
        dname = r.designer or '未指定'
        if dname not in designer_stats:
            designer_stats[dname] = {'total': 0, 'completed': 0, 'items': []}
        designer_stats[dname]['total'] += 1
        if r.completion_status == '已完成':
            designer_stats[dname]['completed'] += 1
        designer_stats[dname]['items'].append({
            'projectName': r.project_name or '',
            'type': '施工图',
            'content': r.drawing_content or '',
            'completionStatus': r.completion_status or '',
            'endDate': r.end_date or '',
        })

    for r in technical_records:
        dname = r.designer or '未指定'
        if dname not in designer_stats:
            designer_stats[dname] = {'total': 0, 'completed': 0, 'items': []}
        designer_stats[dname]['total'] += 1
        designer_stats[dname]['items'].append({
            'projectName': '',
            'type': '技术要求',
            'content': r.equipment_name or '',
            'completionStatus': '',
            'endDate': r.design_complete_date or '',
        })

    sorted_designers = sorted(designer_stats.items(), key=lambda x: x[1]['total'], reverse=True)
    labels = [d[0] for d in sorted_designers]
    details = []
    for dname in labels:
        s = designer_stats[dname]
        details.append({
            'designer': dname,
            'total': s['total'],
            'completed': s['completed'],
            'items': s['items'],
        })

    return jsonify({
        'code': 200,
        'data': {
            'labels': labels,
            'totals': [designer_stats[l]['total'] for l in labels],
            'completed': [designer_stats[l]['completed'] for l in labels],
            'details': details,
        }
    })


@app.route('/api/stats/by-time', methods=['GET'])
@login_required
def api_stats_by_time():
    """按时间阶段统计（默认最近12个月，支持按项目过滤）"""
    project = request.args.get('project', '').strip()

    start_str = request.args.get('start_date', '').strip()
    end_str = request.args.get('end_date', '').strip()

    # 解析时间范围
    if start_str and end_str:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({'code': 400, 'msg': '日期格式无效，请使用 YYYY-MM-DD'}), 400
    else:
        # 默认最近12个月
        end_date = datetime.now()
        start_date = end_date - timedelta(days=365)

    # 施工图按月统计（根据end_date计划完成时间）
    cq = ConstructionPlan.query.filter(ConstructionPlan.end_date != '', ConstructionPlan.end_date.isnot(None))
    if project:
        cq = cq.filter(ConstructionPlan.project_name == project)
    construction_records = cq.all()

    # 阶段设计按月统计（根据publish_date）
    pq = PhasePlan.query.filter(PhasePlan.publish_date != '', PhasePlan.publish_date.isnot(None))
    if project:
        pq = pq.filter(PhasePlan.project_name == project)
    phase_records = pq.all()

    # 技术要求按月统计（根据design_complete_date）
    tq = TechnicalPlan.query.filter(TechnicalPlan.design_complete_date != '', TechnicalPlan.design_complete_date.isnot(None))
    technical_records = tq.all()

    # 生成月份标签
    monthly_data = {}
    current = start_date.replace(day=1)
    while current <= end_date:
        key = current.strftime('%Y-%m')
        monthly_data[key] = {'construction': 0, 'phase': 0, 'technical': 0, 'completed': 0}
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    for r in construction_records:
        if r.end_date:
            try:
                dt = datetime.strptime(r.end_date, '%Y-%m-%d')
                if start_date <= dt <= end_date:
                    key = dt.strftime('%Y-%m')
                    if key in monthly_data:
                        monthly_data[key]['construction'] += 1
                        if r.completion_status == '已完成':
                            monthly_data[key]['completed'] += 1
            except ValueError:
                pass

    for r in phase_records:
        if r.publish_date:
            try:
                dt = datetime.strptime(r.publish_date, '%Y-%m-%d')
                if start_date <= dt <= end_date:
                    key = dt.strftime('%Y-%m')
                    if key in monthly_data:
                        monthly_data[key]['phase'] += 1
            except ValueError:
                pass

    for r in technical_records:
        if r.design_complete_date:
            try:
                dt = datetime.strptime(r.design_complete_date, '%Y-%m-%d')
                if start_date <= dt <= end_date:
                    key = dt.strftime('%Y-%m')
                    if key in monthly_data:
                        monthly_data[key]['technical'] += 1
            except ValueError:
                pass

    labels = list(monthly_data.keys())
    totals = [monthly_data[k]['construction'] for k in labels]
    completed = [monthly_data[k]['completed'] for k in labels]
    phase_totals = [monthly_data[k]['phase'] for k in labels]
    technical_totals = [monthly_data[k]['technical'] for k in labels]

    return jsonify({
        'code': 200,
        'data': {
            'labels': labels,
            'totals': totals,
            'completed': completed,
            'phaseTotals': phase_totals,
            'technicalTotals': technical_totals
        }
    })


@app.route('/api/stats/filter-options', methods=['GET'])
@login_required
def api_stats_filter_options():
    """返回所有可用的筛选选项（用于前端下拉框）"""
    # 从施工图表提取
    construction_records = ConstructionPlan.query.all()
    # 从阶段设计表提取
    phase_records = PhasePlan.query.all()
    # 从技术要求表提取
    technical_records = TechnicalPlan.query.all()

    projects = set()
    majors = set()
    designers = set()
    unit_projects = set()
    phases = set()

    for r in construction_records:
        if r.project_name:
            projects.add(r.project_name)
        if r.major:
            majors.add(r.major)
        if r.designer:
            designers.add(r.designer)
        if r.unit_project:
            unit_projects.add(r.unit_project)

    for r in phase_records:
        if r.project_name:
            projects.add(r.project_name)
        if r.design_phase:
            phases.add(r.design_phase)

    for r in technical_records:
        if r.designer:
            designers.add(r.designer)

    return jsonify({
        'code': 200,
        'data': {
            'projects': sorted(list(projects)),
            'majors': sorted(list(majors)),
            'designers': sorted(list(designers)),
            'unitProjects': sorted(list(unit_projects)),
            'phases': sorted(list(phases)),
        }
    })


@app.route('/api/dashboard', methods=['GET'])
@login_required
def api_dashboard():
    """项目驾驶舱综合数据（支持按项目筛选）"""
    project_name = request.args.get('project', '').strip()
    
    # 构建查询条件
    query = ConstructionPlan.query
    if project_name:
        query = query.filter(ConstructionPlan.project_name == project_name)
    
    all_construction = query.all()
    
    # 只分析施工图计划数据
    construction_total = len(all_construction)
    construction_completed = sum(1 for r in all_construction if r.completion_status == '已完成')
    construction_ongoing = sum(1 for r in all_construction if r.completion_status == '进行中')
    construction_not_started = sum(1 for r in all_construction if r.completion_status == '未开始')
    construction_delayed = sum(1 for r in all_construction if r.completion_status == '已延期')

    projects_data = {}
    project_major_count = {}  # 热力图数据
    for r in all_construction:
        pname = r.project_name or '未定义项目'
        if pname not in projects_data:
            projects_data[pname] = {
                'total': 0, 'completed': 0, 'ongoing': 0,
                'notStarted': 0, 'delayed': 0
            }
        projects_data[pname]['total'] += 1
        if r.completion_status == '已完成':
            projects_data[pname]['completed'] += 1
        elif r.completion_status == '进行中':
            projects_data[pname]['ongoing'] += 1
        elif r.completion_status == '未开始':
            projects_data[pname]['notStarted'] += 1
        elif r.completion_status == '已延期':
            projects_data[pname]['delayed'] += 1
        
        # 热力图统计
        major = r.major or '未分类'
        key = f"{pname}|{major}"
        project_major_count[key] = project_major_count.get(key, 0) + 1

    projects_summary = []
    for pname, pdata in projects_data.items():
        completion_rate = round(pdata['completed'] / pdata['total'] * 100, 1) if pdata['total'] > 0 else 0
        delay_rate = round(pdata['delayed'] / pdata['total'] * 100, 1) if pdata['total'] > 0 else 0
        risk_level = 'high' if delay_rate > 20 else ('medium' if delay_rate > 10 else 'low')
        projects_summary.append({
            'name': pname, 'total': pdata['total'], 'completed': pdata['completed'],
            'ongoing': pdata['ongoing'], 'notStarted': pdata['notStarted'], 'delayed': pdata['delayed'],
            'completionRate': completion_rate, 'delayRate': delay_rate, 'riskLevel': risk_level
        })

    major_stats = {}
    for r in all_construction:
        major = r.major or '未分类'
        if major not in major_stats:
            major_stats[major] = {'total': 0, 'completed': 0, 'delayed': 0}
        major_stats[major]['total'] += 1
        if r.completion_status == '已完成':
            major_stats[major]['completed'] += 1
        elif r.completion_status == '已延期':
            major_stats[major]['delayed'] += 1

    major_comparison = []
    for major, mdata in major_stats.items():
        completion_rate = round(mdata['completed'] / mdata['total'] * 100, 1) if mdata['total'] > 0 else 0
        major_comparison.append({'major': major, 'total': mdata['total'], 'completed': mdata['completed'], 'completionRate': completion_rate})
    major_comparison.sort(key=lambda x: x['completionRate'], reverse=True)

    designer_stats = {}
    for r in all_construction:
        dname = r.designer or '未指定'
        if dname not in designer_stats:
            designer_stats[dname] = {'total': 0, 'completed': 0, 'delayed': 0, 'ongoing': 0}
        designer_stats[dname]['total'] += 1
        if r.completion_status == '已完成':
            designer_stats[dname]['completed'] += 1
        elif r.completion_status == '已延期':
            designer_stats[dname]['delayed'] += 1
        elif r.completion_status == '进行中':
            designer_stats[dname]['ongoing'] += 1

    designer_workload = []
    for dname, ddata in designer_stats.items():
        completion_rate = round(ddata['completed'] / ddata['total'] * 100, 1) if ddata['total'] > 0 else 0
        delay_rate = round(ddata['delayed'] / ddata['total'] * 100, 1) if ddata['total'] > 0 else 0
        designer_workload.append({
            'designer': dname, 'total': ddata['total'], 'completed': ddata['completed'],
            'ongoing': ddata['ongoing'], 'delayed': ddata['delayed'],
            'completionRate': completion_rate, 'delayRate': delay_rate
        })
    designer_workload.sort(key=lambda x: x['total'], reverse=True)

    # 延期排行榜（优化：使用字典）
    delay_counts = {}
    for r in all_construction:
        if r.completion_status == '已延期':
            dname = r.designer or '未指定'
            delay_counts[dname] = delay_counts.get(dname, 0) + 1
    delay_ranking = [{'designer': dname, 'count': cnt} for dname, cnt in delay_counts.items()]
    delay_ranking.sort(key=lambda x: x['count'], reverse=True)

    # 热力图数据（从预计算字典生成）
    heatmap_data = []
    all_projects = list(projects_data.keys())
    all_majors = list(major_stats.keys())
    for key, count in project_major_count.items():
        parts = key.split('|')
        heatmap_data.append({'project': parts[0], 'major': parts[1], 'count': count})

    # 获取所有项目列表（用于筛选） - 只从施工图计划中提取
    all_projects_list = sorted([p[0] for p in ConstructionPlan.query.with_entities(ConstructionPlan.project_name).distinct() if p[0]])

    return jsonify({
        'code': 200,
        'data': {
            'overview': {
                'constructionTotal': construction_total, 'constructionCompleted': construction_completed,
                'constructionOngoing': construction_ongoing, 'constructionNotStarted': construction_not_started,
                'constructionDelayed': construction_delayed,
                'completionRate': round(construction_completed / construction_total * 100, 1) if construction_total > 0 else 0,
                'delayRate': round(construction_delayed / construction_total * 100, 1) if construction_total > 0 else 0
            },
            'projectProgress': projects_summary,
            'majorComparison': major_comparison,
            'designerWorkload': designer_workload[:15],
            'delayRanking': delay_ranking[:10],
            'heatmapData': heatmap_data,
            'allProjects': all_projects,
            'allMajors': all_majors,
            'projectFilterOptions': all_projects_list,
            'selectedProject': project_name
        }
    })


# ==================== Workload Statistics API ====================

@app.route('/api/project-names', methods=['GET'])
@login_required
def api_get_project_names():
    """获取所有项目名称（用于下拉选择）"""
    keyword = request.args.get('keyword', '').strip()
    
    # 仅从项目名称字典表获取
    query = ProjectName.query
    if keyword:
        query = query.filter(ProjectName.name.contains(keyword))
    
    records = query.order_by(ProjectName.name).all()
    
    return jsonify({
        'code': 200,
        'data': [r.to_dict() for r in records]
    })


@app.route('/api/project-names', methods=['POST'])
@edit_required
def api_add_project_name():
    """新增项目名称"""
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'code': 400, 'msg': '项目名称不能为空'}), 400
    
    name = data['name'].strip()
    if ProjectName.query.filter_by(name=name).first():
        return jsonify({'code': 400, 'msg': '项目名称已存在'}), 400
    
    project_name = ProjectName(name=name)
    db.session.add(project_name)
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '添加成功', 'data': project_name.to_dict()})


@app.route('/api/project-names/<int:name_id>', methods=['DELETE'])
@edit_required
def api_delete_project_name(name_id):
    """删除项目名称"""
    project_name = db.session.get(ProjectName, name_id)
    if not project_name:
        return jsonify({'code': 404, 'msg': '项目名称不存在'}), 404
    
    db.session.delete(project_name)
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '删除成功'})


@app.route('/api/project-names/<int:name_id>', methods=['PUT'])
@edit_required
def api_update_project_name(name_id):
    """编辑项目名称"""
    project_name = db.session.get(ProjectName, name_id)
    if not project_name:
        return jsonify({'code': 404, 'msg': '项目名称不存在'}), 404
    
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'code': 400, 'msg': '项目名称不能为空'}), 400
    
    new_name = data['name'].strip()
    existing = ProjectName.query.filter(ProjectName.id != name_id, ProjectName.name == new_name).first()
    if existing:
        return jsonify({'code': 400, 'msg': '项目名称已存在'}), 400
    
    project_name.name = new_name
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '更新成功', 'data': project_name.to_dict()})


@app.route('/api/workload', methods=['GET'])
@login_required
def api_get_workload():
    """获取工作量统计数据"""
    keyword = request.args.get('keyword', '').strip()
    month = request.args.get('month', '').strip()
    major = request.args.get('major', '').strip()
    name = request.args.get('name', '').strip()
    project = request.args.get('project', '').strip()
    
    query = WorkloadStats.query
    
    if month:
        query = query.filter(WorkloadStats.month == month)
    if major:
        query = query.filter(WorkloadStats.major == major)
    if name:
        query = query.filter(WorkloadStats.name == name)
    if project:
        query = query.filter(WorkloadStats.project_name == project)
    if keyword:
        query = query.filter(db.or_(
            WorkloadStats.project_name.contains(keyword),
            WorkloadStats.work_content.contains(keyword),
            WorkloadStats.name.contains(keyword)
        ))
    
    records = query.order_by(WorkloadStats.month.desc(), WorkloadStats.id.desc()).all()
    
    return jsonify({
        'code': 200,
        'data': [r.to_dict() for r in records]
    })


@app.route('/api/workload', methods=['POST'])
@login_required
def api_add_workload():
    """新增工作量记录"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400
    
    required_fields = ['month', 'name', 'projectName']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'code': 400, 'msg': f'{field} 不能为空'}), 400
    
    workload = WorkloadStats(
        month=data['month'],
        department=data.get('department'),
        name=data['name'],
        major=data.get('major'),
        major_director=data.get('majorDirector'),
        role=data.get('role'),
        work_content=data.get('workContent'),
        quantitative_index=data.get('quantitativeIndex'),
        project_name=data['projectName'],
        project_manager=data.get('projectManager'),
        design_phase=data.get('designPhase'),
        drawing_name=data.get('drawingName'),
        a1_drawing_count=data.get('a1DrawingCount'),
        manual_page_count=data.get('manualPageCount'),
        completion_rate=data.get('completionRate'),
        actual_work_days=float(data['actualWorkDays']) if data.get('actualWorkDays') else None,
        work_start_date=data.get('workStartDate'),
        work_end_date=data.get('workEndDate'),
        plan_deviation=float(data['planDeviation']) if data.get('planDeviation') else None,
        remarks=data.get('remarks'),
        created_by=session.get('user_id')
    )
    
    db.session.add(workload)
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '添加成功', 'data': workload.to_dict()})


@app.route('/api/workload/<int:record_id>', methods=['PUT'])
@login_required
def api_update_workload(record_id):
    """更新工作量记录"""
    workload = db.session.get(WorkloadStats, record_id)
    if not workload:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400
    
    user = db.session.get(User, session['user_id'])
    # 设计人员只能修改自己创建的未确认记录
    if user.role == 'staff' and (workload.created_by != session['user_id'] or workload.confirmed):
        return jsonify({'code': 403, 'msg': '权限不足'}), 403
    
    if 'month' in data:
        workload.month = data['month']
    if 'department' in data:
        workload.department = data['department']
    if 'name' in data:
        workload.name = data['name']
    if 'major' in data:
        workload.major = data['major']
    if 'majorDirector' in data:
        workload.major_director = data['majorDirector']
    if 'role' in data:
        workload.role = data['role']
    if 'workContent' in data:
        workload.work_content = data['workContent']
    if 'quantitativeIndex' in data:
        workload.quantitative_index = data['quantitativeIndex']
    if 'projectName' in data:
        workload.project_name = data['projectName']
    if 'projectManager' in data:
        workload.project_manager = data['projectManager']
    if 'designPhase' in data:
        workload.design_phase = data['designPhase']
    if 'drawingName' in data:
        workload.drawing_name = data['drawingName']
    if 'a1DrawingCount' in data:
        workload.a1_drawing_count = data['a1DrawingCount']
    if 'manualPageCount' in data:
        workload.manual_page_count = data['manualPageCount']
    if 'completionRate' in data:
        workload.completion_rate = data['completionRate']
    if 'actualWorkDays' in data:
        workload.actual_work_days = float(data['actualWorkDays']) if data['actualWorkDays'] else None
        # 如果修改了实际工作日，重置确认状态
        workload.confirmed = False
        workload.confirmed_by = None
        workload.confirmed_at = None
    if 'workStartDate' in data:
        workload.work_start_date = data['workStartDate']
    if 'workEndDate' in data:
        workload.work_end_date = data['workEndDate']
    if 'planDeviation' in data:
        workload.plan_deviation = float(data['planDeviation']) if data['planDeviation'] else None
    if 'remarks' in data:
        workload.remarks = data['remarks']
    
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '更新成功', 'data': workload.to_dict()})


@app.route('/api/workload/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_workload(record_id):
    """删除工作量记录"""
    workload = db.session.get(WorkloadStats, record_id)
    if not workload:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    
    user = db.session.get(User, session['user_id'])
    # 设计人员只能删除自己创建的未确认记录
    if user.role == 'staff' and (workload.created_by != session['user_id'] or workload.confirmed):
        return jsonify({'code': 403, 'msg': '权限不足'}), 403
    
    db.session.delete(workload)
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '删除成功'})


@app.route('/api/workload/filter-options', methods=['GET'])
@login_required
def api_workload_filter_options():
    """获取工作量统计筛选选项"""
    options = {
        'name': [],
        'major': [],
        'role': [],
        'project_name': [],
        'design_phase': [],
        'work_start_date': [],
        'work_end_date': []
    }
    
    query = WorkloadStats.query
    
    user = db.session.get(User, session['user_id'])
    if user.role == 'staff':
        query = query.filter(WorkloadStats.created_by == session['user_id'])
    
    records = query.all()
    
    for record in records:
        if record.name and record.name not in options['name']:
            options['name'].append(record.name)
        if record.major and record.major not in options['major']:
            options['major'].append(record.major)
        if record.role and record.role not in options['role']:
            options['role'].append(record.role)
        if record.project_name and record.project_name not in options['project_name']:
            options['project_name'].append(record.project_name)
        if record.design_phase and record.design_phase not in options['design_phase']:
            options['design_phase'].append(record.design_phase)
        if record.work_start_date and record.work_start_date not in options['work_start_date']:
            options['work_start_date'].append(record.work_start_date)
        if record.work_end_date and record.work_end_date not in options['work_end_date']:
            options['work_end_date'].append(record.work_end_date)
    
    return jsonify({'code': 200, 'data': options})


@app.route('/api/workload/batch-delete', methods=['POST'])
@login_required
def api_batch_delete_workload():
    """批量删除工作量记录（支持多条件组合）"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求参数无效'}), 400
    
    query = WorkloadStats.query
    
    user = db.session.get(User, session['user_id'])
    if user.role == 'staff':
        query = query.filter(WorkloadStats.created_by == session['user_id'])
    
    if 'name' in data and data['name']:
        query = query.filter(WorkloadStats.name.like(f'%{data["name"]}%'))
    if 'major' in data and data['major']:
        query = query.filter(WorkloadStats.major.like(f'%{data["major"]}%'))
    if 'role' in data and data['role']:
        query = query.filter(WorkloadStats.role.like(f'%{data["role"]}%'))
    if 'project_name' in data and data['project_name']:
        query = query.filter(WorkloadStats.project_name.like(f'%{data["project_name"]}%'))
    if 'design_phase' in data and data['design_phase']:
        query = query.filter(WorkloadStats.design_phase.like(f'%{data["design_phase"]}%'))
    if 'start_date' in data and data['start_date']:
        query = query.filter(WorkloadStats.work_start_date >= data['start_date'])
    if 'end_date' in data and data['end_date']:
        query = query.filter(WorkloadStats.work_end_date <= data['end_date'])
    
    count = query.delete(synchronize_session=False)
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '批量删除成功', 'data': {'count': count}})


@app.route('/api/workload/<int:record_id>/confirm', methods=['POST'])
@edit_required
def api_confirm_workload(record_id):
    """确认实际工作日（专业所长或首席工程师）"""
    workload = db.session.get(WorkloadStats, record_id)
    if not workload:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    
    workload.confirmed = True
    workload.confirmed_by = session.get('user_id')
    workload.confirmed_at = datetime.now()
    
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '确认成功', 'data': workload.to_dict()})


@app.route('/api/workload/<int:record_id>/unconfirm', methods=['POST'])
@edit_required
def api_unconfirm_workload(record_id):
    """取消确认实际工作日"""
    workload = db.session.get(WorkloadStats, record_id)
    if not workload:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    
    workload.confirmed = False
    workload.confirmed_by = None
    workload.confirmed_at = None
    
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '取消确认成功', 'data': workload.to_dict()})


@app.route('/api/export-excel/workload', methods=['GET'])
@login_required
def api_export_workload_excel():
    """导出工作量统计Excel"""
    keyword = request.args.get('keyword', '').strip()
    month = request.args.get('month', '').strip()
    
    query = WorkloadStats.query
    if month:
        query = query.filter(WorkloadStats.month == month)
    if keyword:
        query = query.filter(db.or_(
            WorkloadStats.project_name.contains(keyword),
            WorkloadStats.name.contains(keyword)
        ))
    
    records = query.order_by(WorkloadStats.month.desc(), WorkloadStats.id.desc()).all()
    
    headers = [
        '月份', '所属单位', '姓名', '专业', '专业所长', '人员角色',
        '具体完成工作内容', '量化指标', '项目名称', '项目负责人',
        '设计阶段', '成品套图/文件名称', '折合A1图纸张数', '说明书页数',
        '套图/文件完成率', '实际工作日', '工作开始时间', '工作结束时间',
        '计划偏差(天)', '备注', '确认状态', '确认时间'
    ]
    
    data_rows = []
    for r in records:
        data_rows.append([
            r.month, r.department or '', r.name, r.major or '', r.major_director or '', r.role or '',
            r.work_content or '', r.quantitative_index or '', r.project_name, r.project_manager or '',
            r.design_phase or '', r.drawing_name or '', r.a1_drawing_count or '', r.manual_page_count or '',
            r.completion_rate or '', r.actual_work_days or '', r.work_start_date or '', r.work_end_date or '',
            r.plan_deviation or '', r.remarks or '', '已确认' if r.confirmed else '待确认',
            r.confirmed_at.strftime('%Y-%m-%d %H:%M:%S') if r.confirmed_at else ''
        ])
    
    wb = Workbook()
    ws = wb.active
    ws.title = '工作量统计'
    ws.append(headers)
    
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    for row_data in data_rows:
        ws.append(row_data)
    
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(str(headers[col_idx - 1])) * 2
        for row in data_rows:
            if col_idx <= len(row) and row[col_idx - 1]:
                max_len = max(max_len, len(str(row[col_idx - 1])) * 1.5)
        ws.column_dimensions[col_letter].width = min(max(max_len, 12), 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'工作量统计_{timestamp}.xlsx'
    )


@app.route('/api/import/workload', methods=['POST'])
@login_required
def api_import_workload():
    """导入工作量统计Excel（智能匹配项目名称，未匹配的输出Excel）"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': '请上传文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 400, 'msg': '文件名为空'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'code': 400, 'msg': '仅支持 .xlsx 格式的Excel文件'}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'code': 400, 'msg': f'Excel文件解析失败: {str(e)}'}), 400

    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    if len(rows) < 2:
        return jsonify({'code': 400, 'msg': 'Excel文件没有数据行'}), 400

    header_row = [str(cell).strip() if cell else '' for cell in rows[0]]
    expected_headers = EXCEL_HEADERS.get('workload', [])
    if header_row[:len(expected_headers)] != expected_headers:
        return jsonify({
            'code': 400,
            'msg': f'表头不匹配。期望: {expected_headers}'
        }), 400

    all_project_names = [p.name for p in ProjectName.query.all()]

    def normalize_name(name):
        """标准化名称：去除空格、特殊字符，转为小写"""
        if not name:
            return ''
        import re
        return re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', name).lower().strip()

    def levenshtein_distance(s1, s2):
        """计算Levenshtein距离"""
        if len(s1) < len(s2):
            return levenshtein_distance(s2, s1)
        if len(s2) == 0:
            return len(s1)
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        return previous_row[-1]

    def smart_match_project(project_name):
        """智能匹配项目名称"""
        if not project_name or not project_name.strip():
            return None, False
        
        normalized_input = normalize_name(project_name)
        if not normalized_input:
            return None, False

        for name in all_project_names:
            if name == project_name:
                return name, True
            
            normalized_name = normalize_name(name)
            if normalized_name == normalized_input:
                return name, True
            
            if normalized_input in normalized_name or normalized_name in normalized_input:
                return name, True
            
            distance = levenshtein_distance(normalized_input, normalized_name)
            max_len = max(len(normalized_input), len(normalized_name))
            if max_len > 0 and distance / max_len < 0.2:
                return name, True

        return None, False

    matched_data = []
    unmatched_rows = []
    total_rows = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            values = [str(cell).strip() if cell is not None else '' for cell in row]
            
            if not values[0] or not values[2] or not values[8]:
                continue

            total_rows += 1
            project_name = values[8]
            matched_name, is_matched = smart_match_project(project_name)
            
            if is_matched:
                matched_data.append({
                    'values': values,
                    'project_name': matched_name
                })
            else:
                best_match = find_best_match(project_name, all_project_names)
                unmatched_rows.append({
                    'values': values,
                    'original_project_name': project_name,
                    'suggested_name': best_match or ''
                })

        except Exception as e:
            continue

    # 导入匹配的数据
    success_count = 0
    update_count = 0
    fail_count = 0
    batch_size = 50
    batch_count = 0

    def build_unique_key(values, project_name):
        """构建唯一标识键（用于判断是否重复）"""
        return (
            values[2],      # 姓名
            values[3] or '', # 专业
            values[5] or '', # 人员角色
            project_name,    # 项目名称（使用匹配后的）
            values[10] or '', # 设计阶段
            values[11] or '', # 成品套图/文件名称
            values[16] or '', # 工作开始时间
            values[17] or ''  # 工作结束时间
        )

    existing_records = {}
    for record in WorkloadStats.query.all():
        key = (
            record.name,
            record.major or '',
            record.role or '',
            record.project_name,
            record.design_phase or '',
            record.drawing_name or '',
            record.work_start_date or '',
            record.work_end_date or ''
        )
        existing_records[key] = record

    for item in matched_data:
        try:
            values = item['values']
            project_name = item['project_name']
            unique_key = build_unique_key(values, project_name)
            
            if unique_key in existing_records:
                existing = existing_records[unique_key]
                existing.month = values[0]
                existing.department = values[1] if values[1] else None
                existing.major_director = values[4] if values[4] else None
                existing.work_content = values[6] if values[6] else None
                existing.quantitative_index = values[7] if values[7] else None
                existing.project_manager = values[9] if values[9] else None
                existing.a1_drawing_count = values[12] if values[12] else None
                existing.manual_page_count = values[13] if values[13] else None
                existing.completion_rate = values[14] if values[14] else None
                existing.actual_work_days = float(values[15]) if values[15] else None
                existing.plan_deviation = float(values[18]) if values[18] else None
                existing.remarks = values[19] if values[19] else None
                update_count += 1
            else:
                workload = WorkloadStats(
                    month=values[0],
                    department=values[1] if values[1] else None,
                    name=values[2],
                    major=values[3] if values[3] else None,
                    major_director=values[4] if values[4] else None,
                    role=values[5] if values[5] else None,
                    work_content=values[6] if values[6] else None,
                    quantitative_index=values[7] if values[7] else None,
                    project_name=project_name,
                    project_manager=values[9] if values[9] else None,
                    design_phase=values[10] if values[10] else None,
                    drawing_name=values[11] if values[11] else None,
                    a1_drawing_count=values[12] if values[12] else None,
                    manual_page_count=values[13] if values[13] else None,
                    completion_rate=values[14] if values[14] else None,
                    actual_work_days=float(values[15]) if values[15] else None,
                    work_start_date=values[16] if values[16] else None,
                    work_end_date=values[17] if values[17] else None,
                    plan_deviation=float(values[18]) if values[18] else None,
                    remarks=values[19] if values[19] else None,
                    confirmed=False,
                    created_by=session.get('user_id')
                )
                db.session.add(workload)
                success_count += 1
            
            batch_count += 1
            if batch_count >= batch_size:
                db.session.commit()
                batch_count = 0
                db.session.flush()
                db.session.clear()
        
        except Exception as e:
            db.session.rollback()
            batch_count = 0
            fail_count += 1
    
    if batch_count > 0:
        db.session.commit()

    # 如果有未匹配的行，生成Excel文件返回
    if unmatched_rows:
        # 构建未匹配数据的Excel
        unmatched_headers = expected_headers + ['建议项目名称']
        wb_unmatched = Workbook()
        ws_unmatched = wb_unmatched.active
        ws_unmatched.title = '未匹配数据'
        ws_unmatched.append(unmatched_headers)
        
        from openpyxl.styles import Font, PatternFill, Alignment
        for col_idx, header in enumerate(unmatched_headers, 1):
            cell = ws_unmatched.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        for item in unmatched_rows:
            row_data = item['values'] + [item['suggested_name']]
            ws_unmatched.append(row_data)
        
        # 设置建议项目名称列的背景色
        suggestion_col = len(unmatched_headers)
        for row_idx in range(2, len(unmatched_rows) + 2):
            cell = ws_unmatched.cell(row=row_idx, column=suggestion_col)
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        output_unmatched = io.BytesIO()
        wb_unmatched.save(output_unmatched)
        output_unmatched.seek(0)
        
        import base64
        file_base64 = base64.b64encode(output_unmatched.getvalue()).decode()
        
        return jsonify({
            'code': 200,
            'msg': f'导入完成，成功 {success_count} 条，更新 {update_count} 条，失败 {fail_count} 条，{len(unmatched_rows)} 条未匹配',
            'data': {
                'success': success_count,
                'update': update_count,
                'unmatched': len(unmatched_rows),
                'fail': fail_count,
                'unmatchedFile': file_base64,
                'unmatchedProjects': list(set(item['original_project_name'] for item in unmatched_rows))[:50]
            }
        })

    return jsonify({
        'code': 200,
        'msg': f'导入完成，成功 {success_count} 条，更新 {update_count} 条，失败 {fail_count} 条',
        'data': {
            'success': success_count,
            'update': update_count,
            'unmatched': 0,
            'fail': fail_count
        }
    })


def find_best_match(project_name, all_project_names):
    """找到最相似的项目名称"""
    if not project_name or not all_project_names:
        return None
    
    import re
    def normalize_name(name):
        if not name:
            return ''
        return re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', name).lower().strip()
    
    def levenshtein_distance(s1, s2):
        if len(s1) < len(s2):
            return levenshtein_distance(s2, s1)
        if len(s2) == 0:
            return len(s1)
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        return previous_row[-1]
    
    normalized_input = normalize_name(project_name)
    if not normalized_input:
        return None
    
    best_match = None
    best_similarity = 0
    
    for name in all_project_names:
        normalized_name = normalize_name(name)
        if not normalized_name:
            continue
        
        if normalized_name == normalized_input:
            return name
        
        if normalized_input in normalized_name or normalized_name in normalized_input:
            return name
        
        distance = levenshtein_distance(normalized_input, normalized_name)
        max_len = max(len(normalized_input), len(normalized_name))
        if max_len > 0:
            similarity = 1 - distance / max_len
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = name
    
    # 只有相似度>60%才返回建议
    if best_similarity > 0.6:
        return best_match
    
    return None


# ==================== Error Handlers ====================

@app.errorhandler(404)
def not_found(e):
    return jsonify({'code': 404, 'msg': '资源不存在'}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({'code': 500, 'msg': '服务器内部错误'}), 500


# ==================== Main ====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        init_default_data()
    print('=' * 50)
    print('  设计项目进度管理系统 已启动')
    print('  访问地址: http://localhost:5000')
    print('  默认账号: admin / admin123')
    print('=' * 50)
    app.run(host='0.0.0.0', port=5000, debug=True)
